import csv
import io
import json
import logging
from datetime import datetime, timedelta

import openpyxl
import requests
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import logout, authenticate, login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Sum, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET

logger = logging.getLogger(__name__)

# Optional: ReportLab as fallback if PDF.co fails
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from home.models import (
    Ward, VRA, Clerk, KIEMSKit, Phase, DailyKIEMSEntry, WhatsAppSetting, WhatsAppGroup, Device, DeviceBurnLog
)
from .forms import (
    WardForm, VRAForm, ClerkForm, KIEMSKitForm, PhaseForm,
    DailyKIEMSEntryForm, DailyEntryFilterForm, ImportForm, VenueMappingForm
)


# ==================== HELPER FUNCTIONS ====================

def is_superadmin(user):
    return user.is_superuser


# ==================== AUTH VIEWS ====================

def login_view(request):
    """
    Login view for SuperAdmin panel with enhanced security features
    """
    # Redirect if already logged in
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('superadmin:dashboard')
        else:
            messages.error(request, 'You do not have permission to access the admin panel.')
            logout(request)
            return redirect('superadmin:login')

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        print(email)

        # Validate inputs
        if not email or not password:
            messages.error(request, 'Please provide both email and password.')
            return render(request, 'superadmin/login.html', {'email': email})

        # Attempt authentication
        username = User.objects.filter(email=email).first()

        user = authenticate(request, username=username, password=password)

        if user is not None:
            if user.is_superuser:
                # Login successful
                login(request, user)
                messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
                return redirect('superadmin:dashboard')
            else:
                messages.error(request, 'You do not have administrator privileges.')
        else:
            # Failed login attempt
            messages.error(request, 'Invalid email or password. Please try again.')

        return render(request, 'superadmin/login.html', {'email': email})

    # GET request
    return render(request, 'superadmin/login.html')


def logout_view(request):
    """
    Logout view for SuperAdmin panel
    """
    if request.user.is_authenticated:
        messages.info(request, 'You have been logged out successfully.')
        logout(request)

    return redirect('superadmin:login')


def password_reset_request(request):
    """
    Password reset request view
    """
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()

        if email:
            # Check if user exists with this email
            from django.contrib.auth.models import User
            try:
                user = User.objects.get(email=email)
                if user.is_superuser:
                    # In production, send password reset email
                    messages.success(request, 'Password reset instructions have been sent to your email.')
                    return redirect('superadmin:login')
                else:
                    messages.error(request, 'This email is not associated with an admin account.')
            except User.DoesNotExist:
                messages.error(request, 'No account found with this email address.')
        else:
            messages.error(request, 'Please provide your email address.')

    return render(request, 'superadmin/password_reset.html')


# ==================== DASHBOARD ====================

@login_required
@user_passes_test(is_superadmin)
def dashboard(request):
    """SuperAdmin Dashboard"""
    active_phase = Phase.objects.filter(active=True).first()

    # Key Metrics
    metrics = {
        'total_wards': Ward.objects.count(),
        'total_vras': VRA.objects.filter(active=True).count(),
        'total_clerks': Clerk.objects.filter(active=True).count(),
        'total_kits': KIEMSKit.objects.filter(status=True).count(),
        'inactive_kits': KIEMSKit.objects.filter(status=False).count(),
        'total_phases': Phase.objects.count(),
        'total_entries': DailyKIEMSEntry.objects.count(),
        'total_registered': DailyKIEMSEntry.objects.aggregate(Sum('total_registered'))['total_registered__sum'] or 0,
        'total_transferred': DailyKIEMSEntry.objects.aggregate(Sum('total_transferred'))['total_transferred__sum'] or 0,
        'total_updated': DailyKIEMSEntry.objects.aggregate(Sum('total_updated'))['total_updated__sum'] or 0,
        'venue_mappings': DailyKIEMSEntry.objects.filter(entry_type='VENUE').count(),
        'registration_entries': DailyKIEMSEntry.objects.filter(entry_type='REGISTRATION').count(),
    }

    # Recent entries
    recent_entries = DailyKIEMSEntry.objects.select_related(
        'kiems_kit', 'phase', 'ward', 'vra'
    ).order_by('-created_at')[:10]

    # Today's activity
    today = timezone.now().date()
    today_entries = DailyKIEMSEntry.objects.filter(entry_date=today)
    today_activity = {
        'registered': today_entries.filter(entry_type='REGISTRATION').aggregate(Sum('total_registered'))[
                          'total_registered__sum'] or 0,
        'transferred': today_entries.filter(entry_type='REGISTRATION').aggregate(Sum('total_transferred'))[
                           'total_transferred__sum'] or 0,
        'deleted': today_entries.filter(entry_type='REGISTRATION').aggregate(Sum('total_updated'))[
                       'total_updated__sum'] or 0,
        'venue_mappings': today_entries.filter(entry_type='VENUE').count(),
        'registration_entries': today_entries.filter(entry_type='REGISTRATION').count(),
    }

    # Kits by ward
    kits_by_ward = KIEMSKit.objects.values('ward__name').annotate(
        total=Count('id'),
        active=Count('id', filter=Q(status=True))
    ).order_by('-total')[:5]

    context = {
        'active_phase': active_phase,
        'metrics': metrics,
        'recent_entries': recent_entries,
        'today_activity': today_activity,
        'kits_by_ward': kits_by_ward,
    }
    return render(request, 'superadmin/dashboard.html', context)


# ==================== PHASE CRUD ====================

@login_required
@user_passes_test(is_superadmin)
def phase_list(request):
    """List all phases"""
    phases = Phase.objects.all().order_by('-start_date')
    return render(request, 'superadmin/phase_list.html', {'phases': phases})


@login_required
@user_passes_test(is_superadmin)
def phase_create(request):
    """Create a new phase"""
    if request.method == 'POST':
        form = PhaseForm(request.POST)
        if form.is_valid():
            phase = form.save()
            messages.success(request, f'Phase "{phase.name}" created successfully!')
            return redirect('superadmin:phase_list')
    else:
        form = PhaseForm()
    return render(request, 'superadmin/phase_form.html', {'form': form, 'title': 'Create Phase'})


@login_required
@user_passes_test(is_superadmin)
def phase_edit(request, pk):
    """Edit a phase"""
    phase = get_object_or_404(Phase, pk=pk)
    if request.method == 'POST':
        form = PhaseForm(request.POST, instance=phase)
        if form.is_valid():
            form.save()
            messages.success(request, f'Phase "{phase.name}" updated successfully!')
            return redirect('superadmin:phase_list')
    else:
        form = PhaseForm(instance=phase)
    return render(request, 'superadmin/phase_form.html', {'form': form, 'title': 'Edit Phase'})


@login_required
@user_passes_test(is_superadmin)
@require_POST
def phase_delete(request, pk):
    """Delete a phase"""
    phase = get_object_or_404(Phase, pk=pk)
    phase_name = phase.name
    phase.delete()
    messages.success(request, f'Phase "{phase_name}" deleted successfully!')
    return redirect('superadmin:phase_list')


# ==================== WARD CRUD ====================

@login_required
@user_passes_test(is_superadmin)
def ward_list(request):
    """List all wards"""
    wards = Ward.objects.all().order_by('name')
    return render(request, 'superadmin/ward_list.html', {'wards': wards})


@login_required
@user_passes_test(is_superadmin)
def ward_create(request):
    """Create a new ward"""
    if request.method == 'POST':
        form = WardForm(request.POST)
        if form.is_valid():
            ward = form.save()
            messages.success(request, f'Ward "{ward.name}" created successfully!')
            return redirect('superadmin:ward_list')
    else:
        form = WardForm()
    return render(request, 'superadmin/ward_form.html', {'form': form, 'title': 'Create Ward'})


@login_required
@user_passes_test(is_superadmin)
def ward_edit(request, pk):
    """Edit a ward"""
    ward = get_object_or_404(Ward, pk=pk)
    if request.method == 'POST':
        form = WardForm(request.POST, instance=ward)
        if form.is_valid():
            form.save()
            messages.success(request, f'Ward "{ward.name}" updated successfully!')
            return redirect('superadmin:ward_list')
    else:
        form = WardForm(instance=ward)
    return render(request, 'superadmin/ward_form.html', {'form': form, 'title': 'Edit Ward'})


@login_required
@user_passes_test(is_superadmin)
@require_POST
def ward_delete(request, pk):
    """Delete a ward"""
    ward = get_object_or_404(Ward, pk=pk)
    ward_name = ward.name
    ward.delete()
    messages.success(request, f'Ward "{ward_name}" deleted successfully!')
    return redirect('superadmin:ward_list')


# ==================== STAFF CRUD ====================

@login_required
@user_passes_test(is_superadmin)
def staff_list(request):
    """List all clerks and VRAs in separate tables"""
    # Get all clerks with their related data
    clerks = Clerk.objects.select_related('ward').prefetch_related('kits').all().order_by('ward__name', 'name')

    # Get all VRAs with their related data
    vras = VRA.objects.select_related('ward').all().order_by('ward__name', 'name')

    # Statistics
    stats = {
        'total_clerks': clerks.count(),
        'total_vras': vras.count(),
        'active_clerks': clerks.filter(active=True).count(),
        'active_vras': vras.filter(active=True).count(),
        'total_staff': clerks.count() + vras.count(),
    }

    return render(request, 'superadmin/staff_list.html', {
        'clerks': clerks,
        'vras': vras,
        'stats': stats,
    })


@login_required
@user_passes_test(is_superadmin)
def staff_create(request):
    """Create a new clerk or VRA"""
    staff_type = request.GET.get('type', request.POST.get('type', 'clerk'))

    if request.method == 'POST':
        staff_type = request.POST.get('type', 'clerk')

        # Debug: print what type we're getting
        print(f"Creating staff of type: {staff_type}")

        if staff_type == 'clerk':
            form = ClerkForm(request.POST)
            if form.is_valid():
                clerk = form.save()
                messages.success(request, f'Clerk "{clerk.name}" created successfully!')
                return redirect('superadmin:staff_list')
            else:
                # Print form errors for debugging
                print(f"Clerk form errors: {form.errors}")
        else:  # VRA
            form = VRAForm(request.POST)
            if form.is_valid():
                vra = form.save()
                messages.success(request, f'VRA "{vra.name}" created successfully!')
                return redirect('superadmin:staff_list')
            else:
                # Print form errors for debugging
                print(f"VRA form errors: {form.errors}")
    else:
        # GET request - instantiate the correct form based on type
        if staff_type == 'clerk':
            form = ClerkForm()
        else:
            form = VRAForm()

    return render(request, 'superadmin/staff_form.html', {
        'form': form,
        'staff_type': staff_type,
        'title': f'Create {staff_type.upper()}'
    })


@login_required
@user_passes_test(is_superadmin)
def staff_edit(request, pk, staff_type):
    """Edit a clerk or VRA"""
    if staff_type == 'clerk':
        staff = get_object_or_404(Clerk, pk=pk)
        if request.method == 'POST':
            form = ClerkForm(request.POST, instance=staff)
            if form.is_valid():
                form.save()
                messages.success(request, f'Clerk "{staff.name}" updated successfully!')
                return redirect('superadmin:staff_list')
        else:
            form = ClerkForm(instance=staff)
    else:  # VRA
        staff = get_object_or_404(VRA, pk=pk)
        if request.method == 'POST':
            form = VRAForm(request.POST, instance=staff)
            if form.is_valid():
                form.save()
                messages.success(request, f'VRA "{staff.name}" updated successfully!')
                return redirect('superadmin:staff_list')
        else:
            form = VRAForm(instance=staff)

    return render(request, 'superadmin/staff_form.html', {
        'form': form,
        'staff': staff,
        'staff_type': staff_type,
        'title': f'Edit {staff_type.upper()}'
    })


@login_required
@user_passes_test(is_superadmin)
@require_POST
def staff_delete(request, pk, staff_type):
    """Delete a clerk or VRA"""
    if staff_type == 'clerk':
        staff = get_object_or_404(Clerk, pk=pk)
        staff_name = staff.name
        staff.delete()
        messages.success(request, f'Clerk "{staff_name}" deleted successfully!')
    else:  # VRA
        staff = get_object_or_404(VRA, pk=pk)
        staff_name = staff.name
        staff.delete()
        messages.success(request, f'VRA "{staff_name}" deleted successfully!')

    return redirect('superadmin:staff_list')


@login_required
@user_passes_test(is_superadmin)
def clerk_list_old(request):
    """List all clerks (legacy)"""
    clerks = Clerk.objects.select_related('ward').all().order_by('ward__name', 'name')
    return render(request, 'superadmin/clerk_list.html', {'clerks': clerks})


# ==================== KIEMS KIT CRUD ====================

@login_required
@user_passes_test(is_superadmin)
def kit_list(request):
    """List all KIEMS kits"""
    kits = KIEMSKit.objects.select_related('ward').prefetch_related('assigned_clerks').all().order_by('ward__name',
                                                                                                      'kit_name')
    return render(request, 'superadmin/kit_list.html', {'kits': kits})


@login_required
@user_passes_test(is_superadmin)
def kit_create(request):
    """Create a new KIEMS kit"""
    if request.method == 'POST':
        form = KIEMSKitForm(request.POST)
        if form.is_valid():
            kit = form.save()
            messages.success(request, f'KIEMS Kit "{kit.kit_name}" created successfully!')
            return redirect('superadmin:kit_list')
    else:
        form = KIEMSKitForm()
    return render(request, 'superadmin/kit_form.html', {'form': form, 'title': 'Create KIEMS Kit'})


@login_required
@user_passes_test(is_superadmin)
def kit_edit(request, pk):
    """Edit a KIEMS kit"""
    kit = get_object_or_404(KIEMSKit, pk=pk)
    if request.method == 'POST':
        form = KIEMSKitForm(request.POST, instance=kit)
        if form.is_valid():
            form.save()
            messages.success(request, f'KIEMS Kit "{kit.kit_name}" updated successfully!')
            return redirect('superadmin:kit_list')
    else:
        form = KIEMSKitForm(instance=kit)
    return render(request, 'superadmin/kit_form.html', {'form': form, 'title': 'Edit KIEMS Kit'})


@login_required
@user_passes_test(is_superadmin)
@require_POST
def kit_delete(request, pk):
    """Delete a KIEMS kit"""
    kit = get_object_or_404(KIEMSKit, pk=pk)
    kit_name = kit.kit_name
    kit.delete()
    messages.success(request, f'KIEMS Kit "{kit_name}" deleted successfully!')
    return redirect('superadmin:kit_list')


# ==================== DAILY ENTRY CRUD ====================

@login_required
@user_passes_test(is_superadmin)
def entry_list(request):
    """List all daily entries with filters"""
    entries = DailyKIEMSEntry.objects.select_related(
        'kiems_kit', 'phase', 'ward', 'vra'
    ).all()

    form = DailyEntryFilterForm(request.GET or None)
    filter_params = {}

    if form.is_valid():
        if form.cleaned_data.get('phase'):
            entries = entries.filter(phase=form.cleaned_data['phase'])
            filter_params['phase'] = form.cleaned_data['phase'].id
        if form.cleaned_data.get('ward'):
            entries = entries.filter(ward=form.cleaned_data['ward'])
            filter_params['ward'] = form.cleaned_data['ward'].id
        if form.cleaned_data.get('kit'):
            entries = entries.filter(kiems_kit=form.cleaned_data['kit'])
            filter_params['kit'] = form.cleaned_data['kit'].id
        if form.cleaned_data.get('vra'):
            entries = entries.filter(vra=form.cleaned_data['vra'])
            filter_params['vra'] = form.cleaned_data['vra'].id
        if form.cleaned_data.get('date_from'):
            entries = entries.filter(entry_date__gte=form.cleaned_data['date_from'])
            filter_params['date_from'] = form.cleaned_data['date_from'].isoformat()
        if form.cleaned_data.get('date_to'):
            entries = entries.filter(entry_date__lte=form.cleaned_data['date_to'])
            filter_params['date_to'] = form.cleaned_data['date_to'].isoformat()
        if form.cleaned_data.get('uploaded') == 'True':
            entries = entries.filter(uploaded=True)
            filter_params['uploaded'] = 'True'
        elif form.cleaned_data.get('uploaded') == 'False':
            entries = entries.filter(uploaded=False)
            filter_params['uploaded'] = 'False'
        if form.cleaned_data.get('entry_type'):
            entries = entries.filter(entry_type=form.cleaned_data['entry_type'])
            filter_params['entry_type'] = form.cleaned_data['entry_type']

    # Calculate totals with gender breakdown
    total_registered = entries.aggregate(Sum('total_registered'))['total_registered__sum'] or 0
    total_male = entries.aggregate(Sum('registered_male'))['registered_male__sum'] or 0
    total_female = entries.aggregate(Sum('registered_female'))['registered_female__sum'] or 0
    total_transferred = entries.aggregate(Sum('total_transferred'))['total_transferred__sum'] or 0
    total_updated = entries.aggregate(Sum('total_updated'))['total_updated__sum'] or 0

    # Entry type stats
    entry_type_stats = {
        'venue_count': entries.filter(entry_type='VENUE').count(),
        'registration_count': entries.filter(entry_type='REGISTRATION').count(),
        'venue_registered': entries.filter(entry_type='VENUE').aggregate(Sum('total_registered'))[
                                'total_registered__sum'] or 0,
        'registration_registered': entries.filter(entry_type='REGISTRATION').aggregate(Sum('total_registered'))[
                                       'total_registered__sum'] or 0,
    }

    # Today's statistics with gender breakdown
    today = timezone.now().date()
    today_entries = DailyKIEMSEntry.objects.filter(entry_date=today)
    today_stats = {
        'total_entries': today_entries.count(),
        'unique_kits': today_entries.values('kiems_kit').distinct().count(),
        'total_registered': today_entries.aggregate(Sum('total_registered'))['total_registered__sum'] or 0,
        'registered_male': today_entries.aggregate(Sum('registered_male'))['registered_male__sum'] or 0,
        'registered_female': today_entries.aggregate(Sum('registered_female'))['registered_female__sum'] or 0,
        'total_transferred': today_entries.aggregate(Sum('total_transferred'))['total_transferred__sum'] or 0,
        'total_updated': today_entries.aggregate(Sum('total_updated'))['total_updated__sum'] or 0,
        'venue_mappings': today_entries.filter(entry_type='VENUE').count(),
        'registration_entries': today_entries.filter(entry_type='REGISTRATION').count(),
    }

    paginator = Paginator(entries, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'form': form,
        'is_filtered': bool(request.GET and any(request.GET.values())),
        'total_registered': total_registered,
        'total_male': total_male,
        'total_female': total_female,
        'total_transferred': total_transferred,
        'total_updated': total_updated,
        'today_stats': today_stats,
        'entry_type_stats': entry_type_stats,
        'filter_params': filter_params,
    }
    return render(request, 'superadmin/entry_list.html', context)


@login_required
@user_passes_test(is_superadmin)
def entry_create(request):
    """Create a new daily entry"""
    if request.method == 'POST':
        form = DailyKIEMSEntryForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.entry_type = form.cleaned_data['resolved_entry_type']
            if entry.entry_type == 'REGISTRATION':
                messages.success(request, 'Daily entry created successfully with voter registrations!')
            else:
                messages.success(request, 'Venue mapping created successfully! No WhatsApp notification sent.')
            entry.save()

            # ONLY send WhatsApp notifications for REGISTRATION entries
            if entry.entry_type == 'REGISTRATION':
                print("=" * 50)
                print("📊 REGISTRATION ENTRY CREATED - SENDING WHATSAPP")
                print(f"Ward: {entry.ward.name}")
                print(f"Kit: {entry.kiems_kit.kit_name}")
                print(f"VRA: {entry.vra.name}")
                print(f"Total Registered: {entry.total_registered}")
                print(f"Male: {entry.registered_male}, Female: {entry.registered_female}")
                print("=" * 50)

                try:
                    # Get user settings
                    settings = get_whatsapp_settings(request.user)

                    # Send VRA submission message if enabled
                    if settings and settings.notify_vra:
                        message = format_vra_message(entry)
                        print("📱 Sending VRA submission message...")
                        send_whatsapp_message(message, request.user)

                    # Check if all wards submitted and send grand total
                    print("🔍 Checking if all wards submitted...")
                    check_and_send_daily_report(request.user)

                except Exception as e:
                    print(f"❌ WhatsApp error: {str(e)}")
                    import traceback
                    traceback.print_exc()
            else:
                print("=" * 50)
                print("📍 VENUE MAPPING CREATED - NO WHATSAPP NOTIFICATION")
                print(f"Ward: {entry.ward.name}")
                print(f"Venue: {entry.venue}")
                print(f"Date: {entry.entry_date}")
                print(f"Kit: {entry.kiems_kit.kit_name}")
                print("=" * 50)

            return redirect('superadmin:entry_list')
    else:
        form = DailyKIEMSEntryForm()

    return render(request, 'superadmin/entry_form.html', {
        'form': form,
        'title': 'Create Daily Entry',
        'entry_type': 'REGISTRATION'
    })


@login_required
@user_passes_test(is_superadmin)
def venue_mapping_create(request):
    """Create a venue mapping without registrations"""
    if request.method == 'POST':
        form = VenueMappingForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)

            # Set entry type to VENUE
            entry.entry_type = 'VENUE'
            entry.registered_male = 0
            entry.registered_female = 0
            entry.total_registered = 0
            entry.total_transferred = 0
            entry.total_updated = 0

            entry.save()

            messages.success(request,
                             f'📍 Venue mapping created successfully for {entry.ward.name} on {entry.entry_date}!')

            print("=" * 50)
            print("📍 VENUE MAPPING CREATED - NO WHATSAPP NOTIFICATION")
            print(f"Ward: {entry.ward.name}")
            print(f"Venue: {entry.venue}")
            print(f"Date: {entry.entry_date}")
            print(f"Kit: {entry.kiems_kit.kit_name}")
            print("=" * 50)

            return redirect('superadmin:entry_list')
    else:
        form = VenueMappingForm()

    return render(request, 'superadmin/venue_mapping_form.html', {
        'form': form,
        'title': 'Create Venue Mapping'
    })


@login_required
@user_passes_test(is_superadmin)
def entry_edit(request, pk):
    """Edit a daily entry"""
    entry = get_object_or_404(DailyKIEMSEntry, pk=pk)

    if request.method == 'POST':
        form = DailyKIEMSEntryForm(request.POST, instance=entry)
        if form.is_valid():
            updated_entry = form.save(commit=False)

            # Determine entry type based on registrations
            has_registrations = (updated_entry.registered_male > 0 or updated_entry.registered_female > 0)

            if has_registrations:
                updated_entry.entry_type = 'REGISTRATION'
                messages.success(request, '✅ Daily entry updated successfully with voter registrations!')
            else:
                updated_entry.entry_type = 'VENUE'
                messages.success(request, '📍 Venue mapping updated successfully!')

            updated_entry.save()

            # ONLY send WhatsApp notifications for REGISTRATION entries
            if updated_entry.entry_type == 'REGISTRATION':
                try:
                    settings = get_whatsapp_settings(request.user)
                    if settings and settings.notify_edit:
                        message = format_vra_update_message(updated_entry)
                        send_whatsapp_message(message, request.user)
                except Exception as e:
                    print(f"WhatsApp error: {str(e)}")
            else:
                print(f"✏️ EDIT SKIPPED - Entry {pk} is VENUE type, no WhatsApp notification")

            return redirect('superadmin:entry_list')
    else:
        form = DailyKIEMSEntryForm(instance=entry)

    return render(request, 'superadmin/entry_form.html', {
        'form': form,
        'title': 'Edit Daily Entry',
        'entry': entry,
        'entry_type': entry.entry_type
    })


@login_required
@user_passes_test(is_superadmin)
@require_POST
def entry_delete(request, pk):
    """Delete a daily entry"""
    entry = get_object_or_404(DailyKIEMSEntry, pk=pk)
    entry_type = entry.entry_type
    entry.delete()

    if entry_type == 'VENUE':
        messages.success(request, '📍 Venue mapping deleted successfully!')
    else:
        messages.success(request, '📊 Daily entry deleted successfully!')

    return redirect('superadmin:entry_list')


# ==================== IMPORT/EXPORT ====================

@login_required
@user_passes_test(is_superadmin)
@require_GET
def export_data(request):
    """Export data in CSV or Excel format with filters"""
    # Get parameters from request
    model_type = request.GET.get('model_type', 'entry')
    export_format = request.GET.get('format', 'csv')

    # Validate model_type
    valid_models = ['ward', 'vra', 'clerk', 'kiemskit', 'phase', 'entry']
    if model_type not in valid_models:
        messages.error(request, f'Invalid model type: {model_type}')
        return redirect('superadmin:entry_list')

    # Validate format
    if export_format not in ['csv', 'xlsx']:
        messages.error(request, f'Invalid format: {export_format}')
        return redirect('superadmin:entry_list')

    # Get filter parameters from request
    filter_params = {}
    if request.GET.get('phase'):
        filter_params['phase'] = request.GET.get('phase')
    if request.GET.get('ward'):
        filter_params['ward'] = request.GET.get('ward')
    if request.GET.get('kit'):
        filter_params['kit'] = request.GET.get('kit')
    if request.GET.get('vra'):
        filter_params['vra'] = request.GET.get('vra')
    if request.GET.get('date_from'):
        filter_params['date_from'] = request.GET.get('date_from')
    if request.GET.get('date_to'):
        filter_params['date_to'] = request.GET.get('date_to')
    if request.GET.get('uploaded'):
        filter_params['uploaded'] = request.GET.get('uploaded')
    if request.GET.get('entry_type'):
        filter_params['entry_type'] = request.GET.get('entry_type')

    timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    filename = f"export_{model_type}_{timestamp}"

    # Map model types to data with filter support
    model_map = {
        'ward': {
            'queryset': Ward.objects.all(),
            'headers': ['Name', 'Code'],
            'rows': lambda q: [[w.name, w.code] for w in q]
        },
        'vra': {
            'queryset': VRA.objects.select_related('ward').all(),
            'headers': ['Name', 'Ward', 'Active', 'Device Token'],
            'rows': lambda q: [[v.name, v.ward.name, v.active, v.device_token] for v in q]
        },
        'clerk': {
            'queryset': Clerk.objects.select_related('ward').all(),
            'headers': ['Name', 'Ward', 'Active'],
            'rows': lambda q: [[c.name, c.ward.name, c.active] for c in q]
        },
        'kiemskit': {
            'queryset': KIEMSKit.objects.select_related('ward').prefetch_related('assigned_clerks').all(),
            'headers': ['Kit Name', 'Serial No', 'Status', 'Ward', 'Assigned Clerks'],
            'rows': lambda q: [[k.kit_name, k.serial_no, 'Active' if k.status else 'Inactive',
                                k.ward.name, ', '.join([c.name for c in k.assigned_clerks.all()])] for k in q]
        },
        'phase': {
            'queryset': Phase.objects.all(),
            'headers': ['Name', 'Start Date', 'End Date', 'Active'],
            'rows': lambda q: [[p.name, p.start_date, p.end_date, p.active] for p in q]
        },
        'entry': {
            'queryset': DailyKIEMSEntry.objects.select_related(
                'kiems_kit', 'phase', 'ward', 'vra'
            ),
            'headers': [
                'Kit', 'Phase', 'Ward', 'VRA', 'Date', 'Venue',
                'Male', 'Female', 'Total Registered',
                'Transferred', 'Deleted', 'Uploaded', 'Entry Type'
            ],
            'rows': lambda q: [[
                e.kiems_kit.kit_name,
                e.phase.name,
                e.ward.name,
                e.vra.name,
                e.entry_date,
                e.venue,
                e.registered_male,
                e.registered_female,
                e.total_registered,
                e.total_transferred,
                e.total_updated,
                'Yes' if e.uploaded else 'No',
                'Registration' if e.entry_type == 'REGISTRATION' else 'Venue Mapping'
            ] for e in q]
        }
    }

    data = model_map.get(model_type)
    if not data:
        messages.error(request, 'Invalid model type')
        return redirect('superadmin:entry_list')

    # Apply filters for entry model
    queryset = data['queryset']
    if model_type == 'entry' and filter_params:
        if filter_params.get('phase'):
            queryset = queryset.filter(phase_id=filter_params['phase'])
        if filter_params.get('ward'):
            queryset = queryset.filter(ward_id=filter_params['ward'])
        if filter_params.get('kit'):
            queryset = queryset.filter(kiems_kit_id=filter_params['kit'])
        if filter_params.get('vra'):
            queryset = queryset.filter(vra_id=filter_params['vra'])
        if filter_params.get('date_from'):
            queryset = queryset.filter(entry_date__gte=filter_params['date_from'])
        if filter_params.get('date_to'):
            queryset = queryset.filter(entry_date__lte=filter_params['date_to'])
        if filter_params.get('uploaded') == 'True':
            queryset = queryset.filter(uploaded=True)
        elif filter_params.get('uploaded') == 'False':
            queryset = queryset.filter(uploaded=False)
        if filter_params.get('entry_type'):
            queryset = queryset.filter(entry_type=filter_params['entry_type'])

    headers = data['headers']
    rows = data['rows'](queryset)

    # Add filter info to filename
    if filter_params:
        filename += "_filtered"

    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response

    elif export_format == 'xlsx':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        wb.save(response)
        return response

    # Fallback
    messages.error(request, 'Invalid export format')
    return redirect('superadmin:entry_list')


@login_required
@user_passes_test(is_superadmin)
def import_data(request):
    """Import data from CSV or Excel file"""
    if request.method == 'POST':
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            model_type = form.cleaned_data['model_type']

            try:
                # Parse file
                if file.name.endswith('.csv'):
                    decoded = file.read().decode('utf-8')
                    reader = csv.DictReader(io.StringIO(decoded))
                    rows = list(reader)
                elif file.name.endswith(('.xlsx', '.xls')):
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, row)))
                else:
                    messages.error(request, 'Unsupported file format. Please use CSV or Excel.')
                    return redirect('superadmin:import_data')

                success_count = 0
                error_count = 0

                # Import logic based on model type
                if model_type == 'ward':
                    for row in rows:
                        try:
                            Ward.objects.get_or_create(
                                name=row.get('Name') or row.get('name'),
                                defaults={'code': row.get('Code') or row.get('code', '')}
                            )
                            success_count += 1
                        except Exception:
                            error_count += 1

                elif model_type == 'vra':
                    for row in rows:
                        try:
                            ward = Ward.objects.filter(name=row.get('Ward') or row.get('ward')).first()
                            if ward:
                                VRA.objects.get_or_create(
                                    name=row.get('Name') or row.get('name'),
                                    ward=ward,
                                    defaults={
                                        'active': row.get('Active', 'True') in ['True', 'true', '1', 'Yes'],
                                        'device_token': row.get('Device Token') or row.get('device_token', ''),
                                    }
                                )
                                success_count += 1
                            else:
                                error_count += 1
                        except Exception:
                            error_count += 1

                elif model_type == 'clerk':
                    for row in rows:
                        try:
                            ward = Ward.objects.filter(name=row.get('Ward') or row.get('ward')).first()
                            if ward:
                                Clerk.objects.get_or_create(
                                    name=row.get('Name') or row.get('name'),
                                    ward=ward,
                                    defaults={'active': row.get('Active', 'True') in ['True', 'true', '1', 'Yes']}
                                )
                                success_count += 1
                            else:
                                error_count += 1
                        except Exception:
                            error_count += 1

                elif model_type == 'kiemskit':
                    for row in rows:
                        try:
                            ward = Ward.objects.filter(name=row.get('Ward') or row.get('ward')).first()
                            if ward:
                                KIEMSKit.objects.get_or_create(
                                    serial_no=row.get('Serial No') or row.get('serial_no'),
                                    defaults={
                                        'kit_name': row.get('Kit Name') or row.get('kit_name'),
                                        'status': row.get('Status', 'Active') in ['Active', 'active', 'True', 'true',
                                                                                  '1'],
                                        'ward': ward,
                                    }
                                )
                                success_count += 1
                            else:
                                error_count += 1
                        except Exception:
                            error_count += 1

                elif model_type == 'phase':
                    for row in rows:
                        try:
                            Phase.objects.get_or_create(
                                name=row.get('Name') or row.get('name'),
                                defaults={
                                    'start_date': row.get('Start Date') or row.get('start_date'),
                                    'end_date': row.get('End Date') or row.get('end_date'),
                                    'active': row.get('Active', 'True') in ['True', 'true', '1', 'Yes'],
                                }
                            )
                            success_count += 1
                        except Exception:
                            error_count += 1

                elif model_type == 'entry' or model_type == 'venue_mapping':
                    for row in rows:
                        try:
                            kit = KIEMSKit.objects.filter(kit_name=row.get('Kit') or row.get('kit')).first()
                            phase = Phase.objects.filter(name=row.get('Phase') or row.get('phase')).first()
                            ward = Ward.objects.filter(name=row.get('Ward') or row.get('ward')).first()
                            vra = VRA.objects.filter(name=row.get('VRA') or row.get('vra'), ward=ward).first()

                            if kit and phase and ward and vra:
                                # Determine entry type
                                registered_male = int(row.get('Male', 0) or row.get('registered_male', 0))
                                registered_female = int(row.get('Female', 0) or row.get('registered_female', 0))
                                has_registrations = (registered_male > 0 or registered_female > 0)

                                entry_type = 'REGISTRATION' if has_registrations else 'VENUE'

                                DailyKIEMSEntry.objects.get_or_create(
                                    kiems_kit=kit,
                                    phase=phase,
                                    entry_date=row.get('Date') or row.get('entry_date'),
                                    vra=vra,
                                    defaults={
                                        'ward': ward,
                                        'venue': row.get('Venue') or row.get('venue', ''),
                                        'registered_male': registered_male,
                                        'registered_female': registered_female,
                                        'total_registered': registered_male + registered_female,
                                        'total_transferred': int(
                                            row.get('Transferred', 0) or row.get('total_transferred', 0)),
                                        'total_updated': int(row.get('Deleted', 0) or row.get('total_updated', 0)),
                                        'uploaded': row.get('Uploaded', 'False') in ['True', 'true', '1', 'Yes'],
                                        'entry_type': entry_type,
                                    }
                                )
                                success_count += 1
                            else:
                                error_count += 1
                        except Exception as e:
                            print(f"Import error: {str(e)}")
                            error_count += 1

                messages.success(request, f'Import completed: {success_count} records imported, {error_count} errors.')
                return redirect('superadmin:dashboard')

            except Exception as e:
                messages.error(request, f'Import failed: {str(e)}')
                return redirect('superadmin:import_data')
    else:
        form = ImportForm()

    return render(request, 'superadmin/import_data.html', {'form': form})


# ==================== CHART DATA API ====================

@login_required
@user_passes_test(is_superadmin)
@require_GET
def get_chart_data(request):
    """API endpoint for chart data"""
    chart_type = request.GET.get('type', 'daily')
    days = int(request.GET.get('days', 30))
    entry_type = request.GET.get('entry_type', 'REGISTRATION')

    data = {'labels': [], 'datasets': []}

    if chart_type == 'daily':
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days)

        daily_data = DailyKIEMSEntry.objects.filter(
            entry_date__gte=start_date,
            entry_date__lte=end_date,
            entry_type=entry_type  # Only include specified entry type
        ).values('entry_date').annotate(
            registered=Sum('total_registered'),
            transferred=Sum('total_transferred'),
            deleted=Sum('total_updated')
        ).order_by('entry_date')

        data['labels'] = [d['entry_date'].strftime('%Y-%m-%d') for d in daily_data]
        data['datasets'] = [
            {
                'label': 'Registered',
                'data': [d['registered'] or 0 for d in daily_data],
                'borderColor': '#34c759',
                'backgroundColor': 'rgba(52, 199, 89, 0.1)',
                'fill': True
            },
            {
                'label': 'Transferred',
                'data': [d['transferred'] or 0 for d in daily_data],
                'borderColor': '#5856d6',
                'backgroundColor': 'rgba(88, 86, 214, 0.1)',
                'fill': True
            },
            {
                'label': 'Deleted',
                'data': [d['deleted'] or 0 for d in daily_data],
                'borderColor': '#ff3b30',
                'backgroundColor': 'rgba(255, 59, 48, 0.1)',
                'fill': True
            }
        ]

    elif chart_type == 'phase':
        phase_data = DailyKIEMSEntry.objects.filter(entry_type=entry_type).values('phase__name').annotate(
            total=Sum('total_registered')
        ).order_by('-total')

        data['labels'] = [d['phase__name'] for d in phase_data]
        data['datasets'] = [
            {
                'label': 'Registrations by Phase',
                'data': [d['total'] or 0 for d in phase_data],
                'backgroundColor': ['#34c759', '#5856d6', '#ff9500', '#ff3b30', '#ff2d55', '#af52de']
            }
        ]

    elif chart_type == 'ward':
        ward_data = DailyKIEMSEntry.objects.filter(entry_type=entry_type).values('ward__name').annotate(
            total=Sum('total_registered')
        ).order_by('-total')[:10]

        data['labels'] = [d['ward__name'] or 'Unknown' for d in ward_data]
        data['datasets'] = [
            {
                'label': 'Registrations by Ward',
                'data': [d['total'] or 0 for d in ward_data],
                'backgroundColor': 'rgba(144, 238, 144, 0.6)',
                'borderColor': '#90EE90',
                'borderWidth': 1
            }
        ]

    elif chart_type == 'kit':
        kit_data = DailyKIEMSEntry.objects.filter(entry_type=entry_type).values('kiems_kit__kit_name').annotate(
            total=Sum('total_registered')
        ).order_by('-total')[:10]

        data['labels'] = [d['kiems_kit__kit_name'] for d in kit_data]
        data['datasets'] = [
            {
                'label': 'Registrations by Kit',
                'data': [d['total'] or 0 for d in kit_data],
                'backgroundColor': ['#34c759', '#5856d6', '#ff9500', '#ff3b30', '#af52de',
                                    '#ff2d55', '#5ac8fa', '#ffcc00', '#ff6b6b', '#6c5ce7']
            }
        ]

    return JsonResponse(data)


@login_required
@user_passes_test(is_superadmin)
@require_GET
def get_clerks_by_ward(request):
    """API endpoint to get clerks for a specific ward"""
    ward_id = request.GET.get('ward_id')
    if not ward_id:
        return JsonResponse({'clerks': [], 'error': 'No ward specified'}, status=400)

    try:
        clerks = Clerk.objects.filter(ward_id=ward_id, active=True).values('id', 'name').order_by('name')
        return JsonResponse({'clerks': list(clerks)})
    except Exception as e:
        return JsonResponse({'clerks': [], 'error': str(e)}, status=500)


@login_required
@user_passes_test(is_superadmin)
@require_GET
def get_vras_by_ward(request):
    """API endpoint to get VRAs for a specific ward"""
    ward_id = request.GET.get('ward_id')
    if not ward_id:
        return JsonResponse({'vras': [], 'error': 'No ward specified'}, status=400)

    try:
        vras = VRA.objects.filter(ward_id=ward_id, active=True).values('id', 'name').order_by('name')
        return JsonResponse({'vras': list(vras)})
    except Exception as e:
        return JsonResponse({'vras': [], 'error': str(e)}, status=500)


# ==================== REPORT GENERATION WITH PDF.CO ====================

def _parse_report_filters(request):
    """Parse and normalize report filter parameters"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    phase_id = request.GET.get('phase')
    ward_id = request.GET.get('ward')
    kit_id = request.GET.get('kit')
    vra_id = request.GET.get('vra')
    uploaded = request.GET.get('uploaded')
    entry_type = request.GET.get('entry_type', 'REGISTRATION')  # Default to REGISTRATION

    today = timezone.now().date()
    default_start = today - timedelta(days=30)

    try:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else default_start
    except (ValueError, TypeError):
        date_from = default_start

    try:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else today
    except (ValueError, TypeError):
        date_to = today

    return {
        'date_from': date_from,
        'date_to': date_to,
        'phase_id': phase_id,
        'ward_id': ward_id,
        'kit_id': kit_id,
        'vra_id': vra_id,
        'uploaded': uploaded,
        'entry_type': entry_type,
    }


def _filter_labels(date_from, date_to, phase_id, ward_id, kit_id, vra_id, uploaded, entry_type):
    """Generate human-readable filter labels"""
    labels = [f"{date_from.strftime('%d %b %Y')} - {date_to.strftime('%d %b %Y')}"]

    if phase_id:
        name = Phase.objects.filter(id=phase_id).values_list('name', flat=True).first()
        if name:
            labels.append(f"Phase: {name}")

    if ward_id:
        name = Ward.objects.filter(id=ward_id).values_list('name', flat=True).first()
        if name:
            labels.append(f"Ward: {name}")

    if kit_id:
        name = KIEMSKit.objects.filter(id=kit_id).values_list('kit_name', flat=True).first()
        if name:
            labels.append(f"Kit: {name}")

    if vra_id:
        name = VRA.objects.filter(id=vra_id).values_list('name', flat=True).first()
        if name:
            labels.append(f"VRA: {name}")

    if uploaded in ('True', 'False'):
        labels.append(f"Status: {'Uploaded' if uploaded == 'True' else 'Pending'}")

    if entry_type:
        labels.append(f"Type: {'Registration' if entry_type == 'REGISTRATION' else 'Venue Mapping'}")

    return labels


def generate_report_html(request):
    """Generate HTML report content"""
    f = _parse_report_filters(request)
    date_from, date_to = f['date_from'], f['date_to']
    phase_id, ward_id, kit_id, vra_id, uploaded, entry_type = (
        f['phase_id'], f['ward_id'], f['kit_id'], f['vra_id'], f['uploaded'], f['entry_type']
    )

    entries = DailyKIEMSEntry.objects.filter(
        entry_date__gte=date_from,
        entry_date__lte=date_to,
        entry_type=entry_type  # Filter by entry type
    ).select_related('kiems_kit', 'phase', 'ward', 'vra').order_by('entry_date')

    if phase_id:
        entries = entries.filter(phase_id=phase_id)
    if ward_id:
        entries = entries.filter(ward_id=ward_id)
    if kit_id:
        entries = entries.filter(kiems_kit_id=kit_id)
    if vra_id:
        entries = entries.filter(vra_id=vra_id)
    if uploaded == 'True':
        entries = entries.filter(uploaded=True)
    elif uploaded == 'False':
        entries = entries.filter(uploaded=False)

    total_registered = entries.aggregate(Sum('total_registered'))['total_registered__sum'] or 0
    total_male = entries.aggregate(Sum('registered_male'))['registered_male__sum'] or 0
    total_female = entries.aggregate(Sum('registered_female'))['registered_female__sum'] or 0
    total_transferred = entries.aggregate(Sum('total_transferred'))['total_transferred__sum'] or 0
    total_entries = entries.count()

    ward_summary = entries.values('ward__name').annotate(
        registered=Sum('total_registered'), male=Sum('registered_male'),
        female=Sum('registered_female'), transferred=Sum('total_transferred'),
        count=Count('id'),
    ).order_by('-registered')

    phase_summary = entries.values('phase__name').annotate(
        registered=Sum('total_registered'), male=Sum('registered_male'),
        female=Sum('registered_female'), transferred=Sum('total_transferred'),
        count=Count('id'),
    ).order_by('-registered')

    kit_summary = entries.values('kiems_kit__kit_name').annotate(
        registered=Sum('total_registered'), male=Sum('registered_male'),
        female=Sum('registered_female'), count=Count('id'),
    ).order_by('-registered')[:20]

    filter_labels = _filter_labels(date_from, date_to, phase_id, ward_id, kit_id, vra_id, uploaded, entry_type)
    scope = " | ".join(filter_labels[1:]) or "All wards, phases and kits"

    entry_type_display = "Registration" if entry_type == 'REGISTRATION' else "Venue Mapping"

    html_string = render_to_string('superadmin/report_template.html', {
        'total_entries': total_entries,
        'total_registered': total_registered,
        'total_male': total_male,
        'total_female': total_female,
        'total_transferred': total_transferred,
        'ward_summary': ward_summary,
        'phase_summary': phase_summary,
        'kit_summary': kit_summary,
        'filter_labels': filter_labels,
        'scope': scope,
        'entry_type_display': entry_type_display,
        'generated_at': timezone.localtime().strftime('%d %b %Y, %H:%M'),
        'date_from': date_from,
        'date_to': date_to,
        'brand_logo_url': getattr(settings, 'BRAND_LOGO_URL', '')
    })

    return HttpResponse(html_string)


@login_required
@user_passes_test(is_superadmin)
def generate_report_pdf(request, as_attachment=False):
    """Generate PDF using PDF.co API"""
    try:
        # Check if API key is configured
        if not hasattr(settings, 'PDF_CO_API_KEY') or not settings.PDF_CO_API_KEY:
            messages.error(request,
                           'PDF.co API key is not configured. Please add PDF_CO_API_KEY to your environment variables.')
            return redirect('superadmin:entry_list')

        # Generate HTML content
        html_response = generate_report_html(request)
        html_string = html_response.content.decode('utf-8')

        # Get filter params for filename
        f = _parse_report_filters(request)
        date_from, date_to = f['date_from'], f['date_to']
        entry_type = f['entry_type']

        # Prepare PDF.co API request
        api_url = f"{getattr(settings, 'PDF_CO_API_URL', 'https://api.pdf.co/v1')}/pdf/convert/from/html"

        payload = json.dumps({
            "name": f"KIEMS_Report_{entry_type}_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.pdf",
            "html": html_string,
            "margin": "20px",
            "paperSize": "Letter",
            "orientation": "Portrait",
            "printBackground": "true",
            "header": "",
            "footer": "",
            "async": False
        })

        headers = {
            'x-api-key': settings.PDF_CO_API_KEY,
            'Content-Type': 'application/json'
        }

        # Call PDF.co API
        response_api = requests.post(api_url, headers=headers, data=payload, timeout=60)

        if response_api.status_code == 200:
            result = response_api.json()

            if result.get('error'):
                messages.error(request, f'PDF generation error: {result["error"]}')
                return redirect('superadmin:entry_list')

            # Check if we have a direct PDF URL
            if result.get('url'):
                # Download the PDF
                pdf_response = requests.get(result['url'], timeout=30)

                if pdf_response.status_code == 200:
                    response = HttpResponse(pdf_response.content, content_type='application/pdf')
                    filename = f"KIEMS_Report_{entry_type}_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.pdf"

                    if as_attachment:
                        response['Content-Disposition'] = f'attachment; filename="{filename}"'
                    else:
                        response['Content-Disposition'] = f'inline; filename="{filename}"'

                    return response
                else:
                    messages.error(request, 'Failed to download generated PDF')
                    return redirect('superadmin:entry_list')
            elif result.get('file'):
                # Some versions return file directly
                response = HttpResponse(result['file'], content_type='application/pdf')
                filename = f"KIEMS_Report_{entry_type}_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.pdf"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            else:
                messages.error(request, 'Unexpected response from PDF service')
                return redirect('superadmin:entry_list')
        else:
            error_msg = response_api.json().get('error', 'Unknown error')
            messages.error(request, f'PDF generation service error: {error_msg}')
            return redirect('superadmin:entry_list')

    except requests.exceptions.Timeout:
        messages.error(request, 'PDF generation timed out. Please try again.')
        return redirect('superadmin:entry_list')
    except requests.exceptions.RequestException as e:
        messages.error(request, f'Network error: {str(e)}')
        return redirect('superadmin:entry_list')
    except Exception as e:
        messages.error(request, f'PDF generation failed: {str(e)}')
        return redirect('superadmin:entry_list')


@login_required
@user_passes_test(is_superadmin)
def generate_report_pdf_fallback(request, as_attachment=False):
    """Fallback PDF generation using ReportLab (local) if PDF.co fails"""
    try:
        if not REPORTLAB_AVAILABLE:
            messages.error(request, 'ReportLab is not available. Please install reportlab or check PDF.co API key.')
            return redirect('superadmin:entry_list')

        f = _parse_report_filters(request)
        date_from, date_to = f['date_from'], f['date_to']
        phase_id, ward_id, kit_id, vra_id, uploaded, entry_type = (
            f['phase_id'], f['ward_id'], f['kit_id'], f['vra_id'], f['uploaded'], f['entry_type']
        )

        # Get data
        entries = DailyKIEMSEntry.objects.filter(
            entry_date__gte=date_from,
            entry_date__lte=date_to,
            entry_type=entry_type
        ).select_related('kiems_kit', 'phase', 'ward', 'vra').order_by('entry_date')

        if phase_id:
            entries = entries.filter(phase_id=phase_id)
        if ward_id:
            entries = entries.filter(ward_id=ward_id)
        if kit_id:
            entries = entries.filter(kiems_kit_id=kit_id)
        if vra_id:
            entries = entries.filter(vra_id=vra_id)
        if uploaded == 'True':
            entries = entries.filter(uploaded=True)
        elif uploaded == 'False':
            entries = entries.filter(uploaded=False)

        total_registered = entries.aggregate(Sum('total_registered'))['total_registered__sum'] or 0
        total_male = entries.aggregate(Sum('registered_male'))['registered_male__sum'] or 0
        total_female = entries.aggregate(Sum('registered_female'))['registered_female__sum'] or 0
        total_transferred = entries.aggregate(Sum('total_transferred'))['total_transferred__sum'] or 0

        response = HttpResponse(content_type='application/pdf')
        entry_type_display = "Registration" if entry_type == 'REGISTRATION' else "Venue Mapping"
        filename = f"KIEMS_Report_{entry_type}_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.pdf"
        if as_attachment:
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
        else:
            response['Content-Disposition'] = f'inline; filename="{filename}"'

        # Build PDF with ReportLab
        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(letter),
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36,
        )

        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Title'],
            fontSize=18,
            textColor=colors.HexColor('#2E7D32'),
            alignment=TA_CENTER,
            spaceAfter=6
        )
        story.append(Paragraph(f"KIEMS {entry_type_display} Report", title_style))

        # Subtitle
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#666666'),
            alignment=TA_CENTER,
            spaceAfter=12
        )
        filter_text = f"{date_from.strftime('%d %b %Y')} - {date_to.strftime('%d %b %Y')} | Type: {entry_type_display}"
        if phase_id:
            phase = Phase.objects.filter(id=phase_id).first()
            if phase: filter_text += f" | Phase: {phase.name}"
        if ward_id:
            ward = Ward.objects.filter(id=ward_id).first()
            if ward: filter_text += f" | Ward: {ward.name}"
        story.append(
            Paragraph(f"Generated: {timezone.localtime().strftime('%d %b %Y, %H:%M')} | {filter_text}", subtitle_style))

        story.append(Spacer(1, 12))

        # Summary stats
        stats_data = [
            ['Total Entries', 'Total Registered', 'Male', 'Female', 'Transferred'],
            [
                str(entries.count()),
                str(total_registered),
                str(total_male),
                str(total_female),
                str(total_transferred)
            ]
        ]

        stats_table = Table(stats_data, colWidths=[100, 100, 80, 80, 80])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5F5')),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 16))

        # Ward Summary
        if entries.exists():
            ward_summary = entries.values('ward__name').annotate(
                registered=Sum('total_registered'),
                male=Sum('registered_male'),
                female=Sum('registered_female'),
                transferred=Sum('total_transferred'),
                count=Count('id'),
            ).order_by('-registered')

            story.append(Paragraph("Summary by Ward", styles['Heading4']))
            story.append(Spacer(1, 8))

            ward_data = [['Ward', 'Entries', 'Male', 'Female', 'Total', 'Transferred']]
            for w in ward_summary:
                ward_data.append([
                    w['ward__name'] or 'Unknown',
                    str(w['count']),
                    str(w['male'] or 0),
                    str(w['female'] or 0),
                    str(w['registered'] or 0),
                    str(w['transferred'] or 0)
                ])

            ward_table = Table(ward_data, colWidths=[120, 60, 60, 60, 60, 70])
            ward_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
            ]))
            story.append(ward_table)
            story.append(Spacer(1, 16))

            # Kit Summary
            kit_summary = entries.values('kiems_kit__kit_name').annotate(
                registered=Sum('total_registered'),
                male=Sum('registered_male'),
                female=Sum('registered_female'),
                count=Count('id'),
            ).order_by('-registered')[:20]

            story.append(Paragraph("Top 20 Kits by Registration", styles['Heading4']))
            story.append(Spacer(1, 8))

            kit_data = [['Kit', 'Entries', 'Male', 'Female', 'Total']]
            for k in kit_summary:
                kit_data.append([
                    k['kiems_kit__kit_name'] or 'Unknown',
                    str(k['count']),
                    str(k['male'] or 0),
                    str(k['female'] or 0),
                    str(k['registered'] or 0)
                ])

            kit_table = Table(kit_data, colWidths=[150, 60, 60, 60, 60])
            kit_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
            ]))
            story.append(kit_table)

        doc.build(story)
        return response

    except Exception as e:
        messages.error(request, f'Fallback PDF generation failed: {str(e)}')
        return redirect('superadmin:entry_list')


@login_required
@user_passes_test(is_superadmin)
def generate_report(request):
    """Generate PDF using PDF.co API with fallback to ReportLab"""
    try:
        # Try PDF.co first
        return generate_report_pdf(request, as_attachment=True)
    except Exception as e:
        # If PDF.co fails, try ReportLab fallback
        messages.warning(request, f'PDF.co service failed, using fallback: {str(e)}')
        return generate_report_pdf_fallback(request, as_attachment=True)


@login_required
@user_passes_test(is_superadmin)
def generate_report_preview(request):
    """Generate HTML preview of the report"""
    return generate_report_html(request)


@login_required
@user_passes_test(is_superadmin)
def generate_report_download(request):
    """Generate and download PDF report"""
    return generate_report(request)


# ==================== KIT REPORT API ====================

@login_required
@user_passes_test(is_superadmin)
def kit_report_api(request, kit_id):
    """API endpoint for kit report data"""
    kit = get_object_or_404(KIEMSKit, id=kit_id)
    entries = DailyKIEMSEntry.objects.filter(kiems_kit=kit).select_related('vra').order_by('-entry_date')

    data = {
        "kit_name": kit.kit_name,
        "serial_no": kit.serial_no,
        "clerks": list(kit.assigned_clerks.values_list('name', flat=True)),
        "vras": list(entries.values_list('vra__name', flat=True).distinct()),
        "entries": [
            {
                "date": e.entry_date.strftime("%b %d, %Y"),
                "vra": e.vra.name,
                "male": e.registered_male,
                "female": e.registered_female,
                "total": e.total_registered,
                "transferred": e.total_transferred,
                "uploaded": e.uploaded,
                "type": "Registration" if e.entry_type == 'REGISTRATION' else "Venue Mapping",
            }
            for e in entries
        ],
        "totals": {
            "male": sum(e.registered_male for e in entries),
            "female": sum(e.registered_female for e in entries),
            "total": sum(e.total_registered for e in entries),
            "transferred": sum(e.total_transferred for e in entries),
        },
    }
    return JsonResponse(data)


# ==================== FILTERED PREVIEW ====================

@login_required
@user_passes_test(is_superadmin)
def filtered_preview(request):
    """Preview filtered entries in a full page"""
    entries = DailyKIEMSEntry.objects.select_related(
        'kiems_kit', 'phase', 'ward', 'vra'
    ).all()

    filter_params = {}
    query_string = ""

    # Apply filters
    if request.GET.get('phase'):
        entries = entries.filter(phase_id=request.GET.get('phase'))
        filter_params['phase'] = request.GET.get('phase')

    if request.GET.get('ward'):
        entries = entries.filter(ward_id=request.GET.get('ward'))
        filter_params['ward'] = request.GET.get('ward')

    if request.GET.get('kit'):
        entries = entries.filter(kiems_kit_id=request.GET.get('kit'))
        filter_params['kit'] = request.GET.get('kit')

    if request.GET.get('vra'):
        entries = entries.filter(vra_id=request.GET.get('vra'))
        filter_params['vra'] = request.GET.get('vra')

    if request.GET.get('date_from'):
        entries = entries.filter(entry_date__gte=request.GET.get('date_from'))
        filter_params['date_from'] = request.GET.get('date_from')

    if request.GET.get('date_to'):
        entries = entries.filter(entry_date__lte=request.GET.get('date_to'))
        filter_params['date_to'] = request.GET.get('date_to')

    if request.GET.get('uploaded') == 'True':
        entries = entries.filter(uploaded=True)
        filter_params['uploaded'] = 'True'
    elif request.GET.get('uploaded') == 'False':
        entries = entries.filter(uploaded=False)
        filter_params['uploaded'] = 'False'

    if request.GET.get('entry_type'):
        entries = entries.filter(entry_type=request.GET.get('entry_type'))
        filter_params['entry_type'] = request.GET.get('entry_type')

    # Build query string for links
    query_string = '&'.join([f'{k}={v}' for k, v in filter_params.items()])

    # Calculate totals
    totals = {
        'registered': entries.aggregate(Sum('total_registered'))['total_registered__sum'] or 0,
        'male': entries.aggregate(Sum('registered_male'))['registered_male__sum'] or 0,
        'female': entries.aggregate(Sum('registered_female'))['registered_female__sum'] or 0,
        'transferred': entries.aggregate(Sum('total_transferred'))['total_transferred__sum'] or 0,
        'updated': entries.aggregate(Sum('total_updated'))['total_updated__sum'] or 0,
    }

    # Build filter summary
    filter_parts = []
    if filter_params.get('phase'):
        phase = Phase.objects.filter(id=filter_params['phase']).first()
        if phase: filter_parts.append(f'Phase: {phase.name}')
    if filter_params.get('ward'):
        ward = Ward.objects.filter(id=filter_params['ward']).first()
        if ward: filter_parts.append(f'Ward: {ward.name}')
    if filter_params.get('kit'):
        kit = KIEMSKit.objects.filter(id=filter_params['kit']).first()
        if kit: filter_parts.append(f'Kit: {kit.kit_name}')
    if filter_params.get('vra'):
        vra = VRA.objects.filter(id=filter_params['vra']).first()
        if vra: filter_parts.append(f'VRA: {vra.name}')
    if filter_params.get('date_from'):
        filter_parts.append(f'From: {filter_params["date_from"]}')
    if filter_params.get('date_to'):
        filter_parts.append(f'To: {filter_params["date_to"]}')
    if filter_params.get('uploaded') == 'True':
        filter_parts.append('Status: Uploaded')
    elif filter_params.get('uploaded') == 'False':
        filter_parts.append('Status: Pending')
    if filter_params.get('entry_type') == 'REGISTRATION':
        filter_parts.append('Type: Registration')
    elif filter_params.get('entry_type') == 'VENUE':
        filter_parts.append('Type: Venue Mapping')

    filter_summary = ' &middot; '.join(filter_parts) if filter_parts else 'No filters applied - showing all entries'

    context = {
        'entries': entries.order_by('-entry_date'),
        'totals': totals,
        'filter_summary': filter_summary,
        'query_string': query_string,
        'generated_at': timezone.localtime().strftime('%d %b %Y, %H:%M'),
    }

    return render(request, 'superadmin/filtered_preview.html', context)


# ==================== PERFORMANCE DASHBOARD ====================

@login_required
@user_passes_test(is_superadmin)
def performance_dashboard(request):
    """Performance dashboard showing best performing kits"""
    period = request.GET.get('period', 'weekly')
    entry_type = request.GET.get('entry_type', 'REGISTRATION')

    # Determine date range
    today = timezone.now().date()
    date_from = None
    date_to = today

    if period == 'daily':
        date_from = today
        period_label = 'Today'
    elif period == 'weekly':
        date_from = today - timedelta(days=7)
        period_label = 'Last 7 Days'
    elif period == 'monthly':
        date_from = today - timedelta(days=30)
        period_label = 'Last 30 Days'
    elif period == 'custom':
        # Handle custom date range
        date_from_str = request.GET.get('date_from')
        date_to_str = request.GET.get('date_to')

        try:
            if date_from_str:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            else:
                date_from = today - timedelta(days=30)

            if date_to_str:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            else:
                date_to = today
        except ValueError:
            date_from = today - timedelta(days=30)
            date_to = today

        period_label = f'{date_from.strftime("%b %d, %Y")} - {date_to.strftime("%b %d, %Y")}'
    else:  # all
        period_label = 'All Time'

    # Build base queryset - only include REGISTRATION entries for performance
    entries = DailyKIEMSEntry.objects.filter(entry_type=entry_type).select_related('kiems_kit', 'ward')

    if date_from:
        entries = entries.filter(entry_date__gte=date_from)
    if date_to and period == 'custom':
        entries = entries.filter(entry_date__lte=date_to)

    # Kit rankings
    kit_rankings = []
    kit_data = entries.values('kiems_kit_id', 'kiems_kit__kit_name', 'kiems_kit__serial_no', 'ward__name').annotate(
        total_registered=Sum('total_registered'),
        total_male=Sum('registered_male'),
        total_female=Sum('registered_female'),
        entry_count=Count('id'),
        total_transferred=Sum('total_transferred'),
    ).order_by('-total_registered')

    total_registered = sum(k['total_registered'] or 0 for k in kit_data)
    max_registered = kit_data[0]['total_registered'] if kit_data else 0

    # Calculate gender totals
    total_male = sum(k['total_male'] or 0 for k in kit_data)
    total_female = sum(k['total_female'] or 0 for k in kit_data)
    gender_total = total_male + total_female

    for kit in kit_data:
        total = kit['total_registered'] or 0
        kit_rankings.append({
            'kit_name': kit['kiems_kit__kit_name'],
            'serial_no': kit['kiems_kit__serial_no'],
            'ward_name': kit['ward__name'],
            'total_registered': total,
            'total_male': kit['total_male'] or 0,
            'total_female': kit['total_female'] or 0,
            'entry_count': kit['entry_count'] or 0,
            'total_transferred': kit['total_transferred'] or 0,
            'pct_of_total': (total / total_registered * 100) if total_registered > 0 else 0,
            'pct_of_max': (total / max_registered * 100) if max_registered > 0 else 0,
            'avg_per_entry': (total / kit['entry_count']) if kit['entry_count'] > 0 else 0,
        })

    # Top kit
    top_kit = kit_rankings[0] if kit_rankings else None

    # Daily performance (last 7 days)
    daily_start = today - timedelta(days=7)
    daily_entries = DailyKIEMSEntry.objects.filter(
        entry_date__gte=daily_start,
        entry_type=entry_type
    ).values('entry_date').annotate(
        entry_count=Count('id'),
        male=Sum('registered_male'),
        female=Sum('registered_female'),
        total_registered=Sum('total_registered'),
        transferred=Sum('total_transferred'),
    ).order_by('-entry_date')

    # Get top kit per day
    daily_performance = []
    for day in daily_entries:
        top_kit_day = DailyKIEMSEntry.objects.filter(
            entry_date=day['entry_date'],
            entry_type=entry_type
        ).values(
            'kiems_kit__kit_name'
        ).annotate(
            total=Sum('total_registered')
        ).order_by('-total').first()

        daily_performance.append({
            'date': day['entry_date'],
            'entry_count': day['entry_count'] or 0,
            'male': day['male'] or 0,
            'female': day['female'] or 0,
            'total_registered': day['total_registered'] or 0,
            'transferred': day['transferred'] or 0,
            'top_kit': top_kit_day['kiems_kit__kit_name'] if top_kit_day else None,
            'top_kit_count': top_kit_day['total'] if top_kit_day else 0,
        })

    # Calculate averages
    total_kits = kit_data.count()
    avg_per_kit = total_registered / total_kits if total_kits > 0 else 0

    # Gender percentages
    gender_male_pct = int((total_male / gender_total * 100)) if gender_total > 0 else 50
    gender_female_pct = int((total_female / gender_total * 100)) if gender_total > 0 else 50

    if gender_total > 0:
        gender_ratio = f"{gender_male_pct}/{gender_female_pct}"
    else:
        gender_ratio = "0/0"

    context = {
        'period': period,
        'period_label': period_label,
        'date_from': date_from if period == 'custom' else None,
        'date_to': date_to if period == 'custom' else None,
        'kit_rankings': kit_rankings,
        'top_kit': top_kit,
        'total_registered': total_registered,
        'total_kits': total_kits,
        'avg_per_kit': int(avg_per_kit),
        'gender_ratio': gender_ratio,
        'gender_male_pct': gender_male_pct,
        'gender_female_pct': gender_female_pct,
        'daily_performance': daily_performance,
        'entry_type': entry_type,
    }

    return render(request, 'superadmin/performance_dashboard.html', context)


# ==================== WHATSAPP BOT INTEGRATION ====================

@login_required
@user_passes_test(is_superadmin)
def whatsapp_status(request):
    """View for WhatsApp integration status page"""
    return render(request, 'superadmin/whatsapp_status.html')


@login_required
@user_passes_test(is_superadmin)
def whatsapp_bot_status(request):
    """Get WhatsApp bot status"""
    try:
        bot_url = getattr(settings, 'WHATSAPP_BOT_URL', 'http://localhost:3000')
        response = requests.get(
            f"{bot_url}/status",
            timeout=5,
            headers={'Content-Type': 'application/json'}
        )

        if response.status_code == 200:
            data = response.json()

            # Handle QR code properly
            qr_data = None
            if data.get('hasQr') and data.get('qr'):
                qr_data = data.get('qr')
                # If QR is not a data URL, it might be raw
                if not qr_data.startswith('data:image'):
                    qr_data = None

            return JsonResponse({
                'success': True,
                'isReady': data.get('isReady', False),
                'status': data.get('status', 'unknown'),
                'hasQr': data.get('hasQr', False),
                'qr': qr_data,
                'uptime': data.get('uptime', 0)
            })
        else:
            return JsonResponse({
                'success': False,
                'isReady': False,
                'status': 'error',
                'error': f'Bot returned status {response.status_code}'
            }, status=503)

    except requests.exceptions.ConnectionError:
        return JsonResponse({
            'success': False,
            'isReady': False,
            'status': 'offline',
            'error': 'WhatsApp bot is not running. Please start the bot server.'
        }, status=503)
    except requests.exceptions.Timeout:
        return JsonResponse({
            'success': False,
            'isReady': False,
            'status': 'timeout',
            'error': 'Connection timeout. Bot server not responding.'
        }, status=503)
    except Exception as e:
        logger.error(f"WhatsApp status error: {str(e)}")
        return JsonResponse({
            'success': False,
            'isReady': False,
            'status': 'error',
            'error': str(e)
        }, status=500)


@login_required
@user_passes_test(is_superadmin)
def whatsapp_groups(request):
    """Get WhatsApp groups and save them to database"""
    try:
        bot_url = getattr(settings, 'WHATSAPP_BOT_URL', 'http://localhost:3000')

        # First check if bot is ready
        status_response = requests.get(
            f"{bot_url}/status",
            timeout=3,
            headers={'Content-Type': 'application/json'}
        )

        if status_response.status_code != 200:
            return JsonResponse({
                'success': False,
                'data': [],
                'error': 'Bot is not responding'
            }, status=503)

        status_data = status_response.json()
        if not status_data.get('isReady', False):
            return JsonResponse({
                'success': False,
                'data': [],
                'error': 'Bot is not ready. Please scan QR code first.'
            }, status=503)

        # Now get groups
        response = requests.get(
            f"{bot_url}/groups",
            timeout=10,
            headers={'Content-Type': 'application/json'}
        )

        if response.status_code == 200:
            data = response.json()
            groups = data.get('data', [])

            # Save each group to database
            saved_groups = []
            for group_data in groups:
                group_id = group_data.get('id')
                group_name = group_data.get('name', f'WhatsApp Group {group_id[:10]}...')

                if group_id:
                    try:
                        # Try to get existing group
                        group, created = WhatsAppGroup.objects.get_or_create(
                            group_id=group_id,
                            defaults={
                                'name': group_name,
                                'is_active': True
                            }
                        )
                        # Update name if it changed
                        if not created and group.name != group_name:
                            group.name = group_name
                            group.save()

                        saved_groups.append({
                            'id': group.group_id,
                            'name': group.name,
                            'participants': group_data.get('participants', 0),
                            'isActive': group.is_active
                        })
                        print(f"? Group saved: {group.name} ({group.group_id})")
                    except Exception as e:
                        print(f"Error saving group {group_id}: {str(e)}")

            return JsonResponse({
                'success': True,
                'data': saved_groups
            })
        else:
            return JsonResponse({
                'success': False,
                'data': [],
                'error': f'Bot returned status {response.status_code}'
            }, status=503)

    except requests.exceptions.ConnectionError:
        return JsonResponse({
            'success': False,
            'data': [],
            'error': 'WhatsApp bot is not running'
        }, status=503)
    except Exception as e:
        logger.error(f"WhatsApp groups error: {str(e)}")
        return JsonResponse({
            'success': False,
            'data': [],
            'error': str(e)
        }, status=500)


@login_required
@user_passes_test(is_superadmin)
@csrf_exempt
def whatsapp_save_settings(request):
    """Save WhatsApp notification settings to database"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        # Parse request data
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()

        group_id = data.get('group_id')
        print(f"? Group ID received: {group_id}")

        notify_vra = data.get('notify_vra', 'true') == 'true'
        notify_edit = data.get('notify_edit', 'true') == 'true'
        notify_daily = data.get('notify_daily', 'true') == 'true'
        notify_grand_total = data.get('notify_grand_total', 'true') == 'true'

        # Get or create WhatsAppSetting for this user
        setting, created = WhatsAppSetting.objects.get_or_create(
            user=request.user,
            defaults={
                'notify_vra': notify_vra,
                'notify_edit': notify_edit,
                'notify_daily': notify_daily,
                'notify_grand_total': notify_grand_total,
            }
        )

        # Update group if provided
        if group_id:
            try:
                # Try to find existing group
                group = WhatsAppGroup.objects.get(group_id=group_id, is_active=True)
                print(f"? Found existing group: {group.name}")
            except WhatsAppGroup.DoesNotExist:
                # Group doesn't exist, create it
                try:
                    # Try to get group name from WhatsApp bot
                    bot_url = getattr(settings, 'WHATSAPP_BOT_URL', 'http://localhost:3000')
                    groups_response = requests.get(
                        f"{bot_url}/groups",
                        timeout=5,
                        headers={'Content-Type': 'application/json'}
                    )
                    group_name = f"WhatsApp Group {group_id[:10]}..."

                    if groups_response.status_code == 200:
                        groups_data = groups_response.json()
                        for g in groups_data.get('data', []):
                            if g.get('id') == group_id:
                                group_name = g.get('name', group_name)
                                break

                    group = WhatsAppGroup.objects.create(
                        group_id=group_id,
                        name=group_name,
                        is_active=True
                    )
                    print(f"? Created new group: {group.name} with ID: {group.group_id}")
                except Exception as e:
                    print(f"? Error creating group: {str(e)}")
                    # Fallback: create with minimal info
                    group = WhatsAppGroup.objects.create(
                        group_id=group_id,
                        name=f"Group {group_id[:15]}...",
                        is_active=True
                    )

            if group:
                setting.default_group = group
                print(f"? Set default_group to: {group.name}")
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Failed to create or find the WhatsApp group'
                }, status=404)

        # Update notification settings
        setting.notify_vra = notify_vra
        setting.notify_edit = notify_edit
        setting.notify_daily = notify_daily
        setting.notify_grand_total = notify_grand_total
        setting.save()

        # Also save in session for quick access
        request.session['whatsapp_settings'] = {
            'group_id': setting.default_group.group_id if setting.default_group else '',
            'notify_vra': setting.notify_vra,
            'notify_edit': setting.notify_edit,
            'notify_daily': setting.notify_daily,
            'notify_grand_total': setting.notify_grand_total,
        }
        request.session.modified = True

        return JsonResponse({
            'success': True,
            'message': 'Settings saved successfully!',
            'data': {
                'group_id': setting.default_group.group_id if setting.default_group else '',
                'group_name': setting.default_group.name if setting.default_group else '',
                'notify_vra': setting.notify_vra,
                'notify_edit': setting.notify_edit,
                'notify_daily': setting.notify_daily,
                'notify_grand_total': setting.notify_grand_total,
            }
        })

    except Exception as e:
        logger.error(f"Error saving WhatsApp settings: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_passes_test(is_superadmin)
@csrf_exempt
def whatsapp_test(request):
    """Send test message via WhatsApp bot"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        # Parse request data
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()

        group_id = data.get('group_id')
        message = data.get('message')

        if not group_id:
            return JsonResponse({
                'success': False,
                'error': 'Please select a group first'
            }, status=400)

        if not message:
            return JsonResponse({
                'success': False,
                'error': 'Message is required'
            }, status=400)

        bot_url = getattr(settings, 'WHATSAPP_BOT_URL', 'http://localhost:3000')
        response = requests.post(
            f"{bot_url}/send",
            json={'groupId': group_id, 'message': message},
            timeout=10,
            headers={'Content-Type': 'application/json'}
        )

        if response.status_code == 200:
            return JsonResponse({
                'success': True,
                'message': 'Test message sent successfully!'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': f'Bot returned status {response.status_code}'
            }, status=503)

    except requests.exceptions.ConnectionError:
        return JsonResponse({
            'success': False,
            'error': 'WhatsApp bot is not running. Please start the bot server.'
        }, status=503)
    except Exception as e:
        logger.error(f"Error sending test message: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_passes_test(is_superadmin)
def whatsapp_get_settings(request):
    """Get WhatsApp notification settings from database"""
    try:
        setting = WhatsAppSetting.objects.filter(user=request.user).first()

        if setting:
            data = {
                'group_id': setting.default_group.group_id if setting.default_group else '',
                'group_name': setting.default_group.name if setting.default_group else '',
                'notify_vra': setting.notify_vra,
                'notify_edit': setting.notify_edit,
                'notify_daily': setting.notify_daily,
                'notify_grand_total': setting.notify_grand_total,
            }
        else:
            data = {
                'group_id': '',
                'group_name': '',
                'notify_vra': True,
                'notify_edit': True,
                'notify_daily': True,
                'notify_grand_total': True,
            }

        return JsonResponse({
            'success': True,
            'data': data
        })
    except Exception as e:
        logger.error(f"Error getting WhatsApp settings: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_passes_test(is_superadmin)
@require_GET
def export_staff(request):
    """Export clerks or VRAs to CSV or Excel"""
    staff_type = request.GET.get('type', 'clerk')
    export_format = request.GET.get('format', 'csv')

    # Validate
    if staff_type not in ['clerk', 'vra']:
        messages.error(request, 'Invalid staff type')
        return redirect('superadmin:staff_list')

    if export_format not in ['csv', 'xlsx']:
        messages.error(request, 'Invalid format')
        return redirect('superadmin:staff_list')

    timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    filename = f"export_{staff_type}s_{timestamp}"

    if staff_type == 'clerk':
        clerks = Clerk.objects.select_related('ward').prefetch_related('kits').all().order_by('ward__name', 'name')
        headers = ['ID', 'Name', 'Ward', 'Assigned Kits', 'Active', 'Created At']
        rows = []
        for clerk in clerks:
            kit_names = ', '.join([kit.kit_name for kit in clerk.kits.all()]) or 'None'
            rows.append([
                clerk.id,
                clerk.name,
                clerk.ward.name if clerk.ward else 'Not Assigned',
                kit_names,
                'Active' if clerk.active else 'Inactive',
                clerk.created_at.strftime('%Y-%m-%d %H:%M') if clerk.created_at else ''
            ])
    else:  # vra
        vras = VRA.objects.select_related('ward').all().order_by('ward__name', 'name')
        headers = ['ID', 'Name', 'Ward', 'Device Token', 'Device Fingerprint', 'Active', 'Created At']
        rows = []
        for vra in vras:
            rows.append([
                vra.id,
                vra.name,
                vra.ward.name if vra.ward else 'Not Assigned',
                vra.device_token or '',
                vra.device_fingerprint or '',
                'Active' if vra.active else 'Inactive',
                vra.created_at.strftime('%Y-%m-%d %H:%M') if vra.created_at else ''
            ])

    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response

    elif export_format == 'xlsx':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = staff_type.capitalize() + 's'
        ws.append(headers)
        for row in rows:
            ws.append(row)
        wb.save(response)
        return response

    messages.error(request, 'Invalid export format')
    return redirect('superadmin:staff_list')


# ==================== WHATSAPP HELPER FUNCTIONS ====================

def get_whatsapp_group(user=None):
    """Get the default WhatsApp group for a user"""
    try:
        if user:
            # Get user's saved group
            setting = WhatsAppSetting.objects.filter(user=user).first()
            if setting and setting.default_group and setting.default_group.is_active:
                return setting.default_group.group_id

        # Fallback: get first active group
        group = WhatsAppGroup.objects.filter(is_active=True).first()
        if group:
            return group.group_id
    except Exception as e:
        print(f"Error getting WhatsApp group: {str(e)}")
    return None


def get_whatsapp_settings(user):
    """Get WhatsApp settings for a user"""
    try:
        setting, created = WhatsAppSetting.objects.get_or_create(
            user=user,
            defaults={
                'notify_vra': True,
                'notify_edit': True,
                'notify_daily': True,
                'notify_grand_total': True,
            }
        )
        return setting
    except Exception as e:
        print(f"Error getting WhatsApp settings: {str(e)}")
        return None


def send_whatsapp_message(message, user=None):
    """Send a message to WhatsApp group"""
    try:
        group_id = get_whatsapp_group(user)
        if not group_id:
            print("No WhatsApp group configured")
            return False

        bot_url = getattr(settings, 'WHATSAPP_BOT_URL', 'http://localhost:3000')
        response = requests.post(
            f"{bot_url}/send",
            json={'groupId': group_id, 'message': message},
            timeout=10,
            headers={'Content-Type': 'application/json'}
        )

        if response.status_code == 200:
            print("? WhatsApp message sent successfully")
            return True
        else:
            print(f"? WhatsApp send failed: {response.text}")
            return False
    except Exception as e:
        print(f"? WhatsApp error: {str(e)}")
        return False


def format_vra_message(entry):
    """Format VRA submission message - Super Simple"""
    message = f"{entry.ward.name.upper()} CONFIRMED ✅ \n"
    message += f"{entry.kiems_kit.kit_name}: MALE:{entry.registered_male} FEMALE:{entry.registered_female} = {entry.total_registered}"
    if entry.total_transferred and entry.total_transferred > 0:
        message += f"\nTransferred: {entry.total_transferred}"
    return message


def format_vra_update_message(entry):
    """Format VRA update message - Super Simple"""
    message = f"{entry.ward.name.upper()} UPDATED ✏️ \n"
    message += f"{entry.kiems_kit.kit_name}: MALE:{entry.registered_male} FEMALE:{entry.registered_female} = {entry.total_registered}"
    if entry.total_transferred and entry.total_transferred > 0:
        message += f"\nTransferred: {entry.total_transferred}"
    return message


def format_grand_total_message(entries, total_wards):
    """Format grand total message when all wards submit - Super Simple"""
    today = timezone.now().date()

    # Calculate grand totals
    total_male = entries.aggregate(Sum('registered_male'))['registered_male__sum'] or 0
    total_female = entries.aggregate(Sum('registered_female'))['registered_female__sum'] or 0
    total_registered = entries.aggregate(Sum('total_registered'))['total_registered__sum'] or 0
    total_transferred = entries.aggregate(Sum('total_transferred'))['total_transferred__sum'] or 0

    # Get per ward breakdown
    ward_data = entries.values('ward__name').annotate(
        male=Sum('registered_male'),
        female=Sum('registered_female'),
        total=Sum('total_registered')
    ).order_by('ward__name')

    # Build message - Simple format
    message = f"📊 DAILY REPORT - {today.strftime('%d %b %Y')}\n"
    message += f"✅ All {total_wards} Wards Submitted!\n\n"

    # Per ward breakdown - one line each
    for w in ward_data:
        message += f"{w['ward__name']}: MALE:{w['male']}  FEMALE:{w['female']} = {w['total']}\n"

    # Grand totals at the end
    message += f"\n📈 TOTAL: {total_male}♂ {total_female}♀ = {total_registered}"
    if total_transferred > 0:
        message += f" | Transferred: {total_transferred}"

    return message


def check_and_send_daily_report(user=None):
    """Check if all wards submitted and send daily report"""
    today = timezone.now().date()

    # Get total wards
    total_wards = Ward.objects.count()
    if total_wards == 0:
        print("No wards found")
        return

    # Count wards that have REGISTRATION entries (not just venue mappings)
    submitted_wards = DailyKIEMSEntry.objects.filter(
        entry_date=today,
        entry_type='REGISTRATION'  # Only count actual registration entries
    ).values('ward').distinct().count()

    print(f"Total wards: {total_wards}, Submitted: {submitted_wards}")

    # If not all wards submitted, return
    if submitted_wards < total_wards:
        print("Not all wards submitted yet")
        return

    print("All wards submitted! Sending grand total...")

    # Get settings for the user
    settings = get_whatsapp_settings(user) if user else None

    # Check if grand total notifications are enabled
    if settings and not settings.notify_grand_total:
        print("Grand total notifications disabled in settings")
        return

    # Get all REGISTRATION entries for today
    entries = DailyKIEMSEntry.objects.filter(
        entry_date=today,
        entry_type='REGISTRATION'  # Only include registration entries
    )

    if not entries.exists():
        print("No registration entries found")
        return

    # Format and send message
    message = format_grand_total_message(entries, total_wards)
    send_whatsapp_message(message, user)


def send_daily_report_job():
    """Job function to send daily report - called from cron or manually"""
    today = timezone.now().date()

    # Get total wards
    total_wards = Ward.objects.count()
    if total_wards == 0:
        print("No wards found")
        return

    # Get REGISTRATION entries for today
    entries = DailyKIEMSEntry.objects.filter(
        entry_date=today,
        entry_type='REGISTRATION'
    )

    if not entries.exists():
        print(f"No registration entries found for {today}")
        return

    # Get admin user for settings
    admin_user = User.objects.filter(is_superuser=True).first()
    if not admin_user:
        print("No admin user found")
        return

    # Get settings
    settings = get_whatsapp_settings(admin_user)

    # Check if daily reports are enabled
    if settings and not settings.notify_daily:
        print("Daily reports disabled")
        return

    # Format and send daily report
    message = format_grand_total_message(entries, total_wards)
    send_whatsapp_message(message, admin_user)
    print(f"Daily report sent for {today}")


@login_required
@user_passes_test(is_superadmin)
def send_daily_report_manual(request):
    """Manually trigger daily report"""
    if request.method == 'POST':
        try:
            send_daily_report_job()
            messages.success(request, 'Daily report sent successfully!')
        except Exception as e:
            messages.error(request, f'Failed to send daily report: {str(e)}')
        return redirect('superadmin:whatsapp_status')

    return redirect('superadmin:whatsapp_status')


# ==================== DEVICE MANAGEMENT ====================

def device_list(request):
    """List all devices with filtering and pagination"""
    devices = Device.objects.select_related('vra', 'vra__ward').all()

    # Filtering
    status_filter = request.GET.get('status', '')
    if status_filter == 'burned':
        devices = devices.filter(is_burned=True)
    elif status_filter == 'unburned':
        devices = devices.filter(is_burned=False)
    elif status_filter == 'active':
        devices = devices.filter(is_active=True)
    elif status_filter == 'inactive':
        devices = devices.filter(is_active=False)

    # Search
    search = request.GET.get('search', '')
    if search:
        devices = devices.filter(
            Q(fingerprint__icontains=search) |
            Q(vra__name__icontains=search) |
            Q(vra__ward__name__icontains=search)
        )

    # Pagination
    paginator = Paginator(devices.order_by('-last_seen'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Statistics
    total_devices = Device.objects.count()
    burned_count = Device.objects.filter(is_burned=True).count()
    unburned_count = Device.objects.filter(is_burned=False).count()
    active_count = Device.objects.filter(is_active=True).count()

    context = {
        'page_obj': page_obj,
        'total_devices': total_devices,
        'burned_count': burned_count,
        'unburned_count': unburned_count,
        'active_count': active_count,
        'search': search,
        'status_filter': status_filter,
        'vras': VRA.objects.filter(active=True).select_related('ward'),
    }
    return render(request, 'superadmin/device_list.html', context)


@login_required
@user_passes_test(is_superadmin)
def device_burn(request):
    """Burn (authorize) a device"""
    if request.method == 'POST':
        fingerprint = request.POST.get('fingerprint')
        vra_id = request.POST.get('vra_id')
        notes = request.POST.get('notes', '')

        if not fingerprint or not vra_id:
            messages.error(request, 'Fingerprint and VRA selection are required.')
            return redirect('superadmin:device_list')

        try:
            device = Device.objects.get(fingerprint=fingerprint)
            vra = VRA.objects.get(id=vra_id)

            # Burn the device
            device.is_burned = True
            device.is_active = True
            device.vra = vra
            device.burn_date = timezone.now()
            device.burn_notes = notes
            device.save()

            # Log the burn
            DeviceBurnLog.objects.create(
                device=device,
                action='BURN',
                performed_by=request.user,
                notes=notes
            )

            # Update VRA's device token
            vra.device_token = fingerprint
            vra.save(update_fields=['device_token'])

            messages.success(request, f'Device burned successfully for VRA: {vra.name}')

        except Device.DoesNotExist:
            messages.error(request, 'Device not found.')
        except VRA.DoesNotExist:
            messages.error(request, 'VRA not found.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    return redirect('superadmin:device_list')


@login_required
@user_passes_test(is_superadmin)
def device_unburn(request, device_id):
    """Unburn (deauthorize) a device"""
    if request.method == 'POST':
        try:
            device = get_object_or_404(Device, id=device_id)

            # Unburn the device
            device.is_burned = False
            device.is_active = False
            device.vra = None
            device.burn_date = None
            device.burn_notes = ''
            device.save()

            # Log the unburn
            DeviceBurnLog.objects.create(
                device=device,
                action='UNBURN',
                performed_by=request.user,
                notes='Unburned by admin'
            )

            messages.success(request, f'Device unburned successfully.')

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    return redirect('superadmin:device_list')


@login_required
@user_passes_test(is_superadmin)
def device_bulk_burn(request):
    """Bulk burn multiple devices"""
    if request.method == 'POST':
        fingerprint_list = request.POST.getlist('fingerprints[]')
        vra_id = request.POST.get('vra_id')
        notes = request.POST.get('notes', '')

        if not fingerprint_list or not vra_id:
            messages.error(request, 'Select devices and a VRA.')
            return redirect('superadmin:device_list')

        try:
            vra = VRA.objects.get(id=vra_id)
            burned_count = 0

            for fingerprint in fingerprint_list:
                try:
                    device = Device.objects.get(fingerprint=fingerprint)
                    if not device.is_burned:
                        device.is_burned = True
                        device.is_active = True
                        device.vra = vra
                        device.burn_date = timezone.now()
                        device.burn_notes = notes
                        device.save()

                        DeviceBurnLog.objects.create(
                            device=device,
                            action='BURN',
                            performed_by=request.user,
                            notes=notes
                        )
                        burned_count += 1
                except Device.DoesNotExist:
                    continue

            messages.success(request, f'Successfully burned {burned_count} devices for VRA: {vra.name}')

        except VRA.DoesNotExist:
            messages.error(request, 'VRA not found.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    return redirect('superadmin:device_list')


@login_required
@user_passes_test(is_superadmin)
def device_logs(request):
    """View device burn logs"""
    logs = DeviceBurnLog.objects.select_related(
        'device', 'performed_by'
    ).all().order_by('-created_at')

    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'total_logs': logs.count(),
    }
    return render(request, 'superadmin/device_logs.html', context)


@login_required
@user_passes_test(is_superadmin)
def api_get_vras_by_ward(request):
    """API endpoint to get VRAs by ward"""
    ward_id = request.GET.get('ward_id')
    if ward_id:
        vras = VRA.objects.filter(ward_id=ward_id, active=True).values('id', 'name')
        return JsonResponse({'vras': list(vras)})
    return JsonResponse({'vras': []})


@login_required
@user_passes_test(is_superadmin)
def api_device_status(request):
    """API endpoint to check device status"""
    fingerprint = request.GET.get('fingerprint')
    if fingerprint:
        try:
            device = Device.objects.select_related('vra', 'vra__ward').get(fingerprint=fingerprint)
            return JsonResponse({
                'exists': True,
                'is_burned': device.is_burned,
                'is_active': device.is_active,
                'vra_id': device.vra_id,
                'vra_name': device.vra.name if device.vra else None,
                'ward_name': device.vra.ward.name if device.vra and device.vra.ward else None,
                'burn_date': device.burn_date.isoformat() if device.burn_date else None,
            })
        except Device.DoesNotExist:
            return JsonResponse({'exists': False})
    return JsonResponse({'error': 'Fingerprint required'}, status=400)


@login_required
@user_passes_test(is_superadmin)
def device_export(request):
    """Export device data as CSV"""
    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="devices_export.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Device ID', 'Fingerprint', 'Platform', 'Screen Resolution',
        'First Seen', 'Last Seen', 'Is Burned', 'Is Active',
        'VRA Name', 'Ward Name', 'Burn Date', 'Burn Notes'
    ])

    devices = Device.objects.select_related('vra', 'vra__ward').all()
    for d in devices:
        writer.writerow([
            d.id,
            d.fingerprint,
            d.platform,
            d.screen_resolution,
            d.first_seen.isoformat() if d.first_seen else '',
            d.last_seen.isoformat() if d.last_seen else '',
            'Yes' if d.is_burned else 'No',
            'Yes' if d.is_active else 'No',
            d.vra.name if d.vra else '',
            d.vra.ward.name if d.vra and d.vra.ward else '',
            d.burn_date.isoformat() if d.burn_date else '',
            d.burn_notes,
        ])

    return response


@login_required
@user_passes_test(is_superadmin)
@require_POST
def authorize_device(request):
    """Authorize (burn) a device"""
    try:
        fingerprint = request.POST.get('fingerprint')
        vra_id = request.POST.get('vra_id')
        notes = request.POST.get('notes', '')

        if not fingerprint:
            messages.error(request, 'Fingerprint required.')
            return redirect('superadmin:device_list')

        device = get_object_or_404(Device, fingerprint=fingerprint)
        vra = get_object_or_404(VRA, id=vra_id) if vra_id else None

        with transaction.atomic():
            device.is_burned = True
            device.is_active = True
            if vra:
                device.vra = vra
                vra.device_token = fingerprint
                vra.save(update_fields=['device_token'])
            device.burn_date = timezone.now()
            device.burn_notes = notes
            device.save()

            DeviceBurnLog.objects.create(
                device=device,
                action='BURN',
                performed_by=request.user,
                notes=notes
            )

        messages.success(request, f'Device authorized successfully!')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')

    return redirect('superadmin:device_list')


@login_required
@user_passes_test(is_superadmin)
@require_POST
def deauthorize_device(request):
    """Deauthorize (unburn) a device"""
    try:
        fingerprint = request.POST.get('fingerprint')
        notes = request.POST.get('notes', '')

        if not fingerprint:
            messages.error(request, 'Fingerprint required.')
            return redirect('superadmin:device_list')

        device = get_object_or_404(Device, fingerprint=fingerprint)

        with transaction.atomic():
            device.is_burned = False
            device.is_active = False

            if device.vra:
                if device.vra.device_token == fingerprint:
                    device.vra.device_token = None
                    device.vra.save(update_fields=['device_token'])
                device.vra = None

            device.burn_date = None
            device.burn_notes = ''
            device.save()

            DeviceBurnLog.objects.create(
                device=device,
                action='UNBURN',
                performed_by=request.user,
                notes=notes
            )

        messages.success(request, 'Device deauthorized successfully!')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')

    return redirect('superadmin:device_list')


@login_required
@user_passes_test(is_superadmin)
@require_POST
def delete_device(request):
    """Delete a device permanently"""
    try:
        fingerprint = request.POST.get('fingerprint')

        if not fingerprint:
            messages.error(request, 'Fingerprint required.')
            return redirect('superadmin:device_list')

        device = get_object_or_404(Device, fingerprint=fingerprint)

        # Delete associated logs first
        DeviceBurnLog.objects.filter(device=device).delete()

        # Remove VRA reference if any
        if device.vra and device.vra.device_token == fingerprint:
            device.vra.device_token = None
            device.vra.save(update_fields=['device_token'])

        device.delete()

        messages.success(request, 'Device deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')

    return redirect('superadmin:device_list')


# ==================== VENUE MANAGEMENT ====================

@login_required
@user_passes_test(is_superadmin)
def venue_management(request):
    """Venue management for daily entries"""
    ward_id = request.GET.get('ward')
    kit_id = request.GET.get('kit')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    entry_type = request.GET.get('entry_type', '')      # '' = all types (fixes the empty-filter bug)
    venue_status = request.GET.get('venue_status', '')  # '', 'set', 'missing'
    search = request.GET.get('search', '')

    base = DailyKIEMSEntry.objects.select_related('kiems_kit', 'ward', 'vra')

    if entry_type in ('VENUE', 'REGISTRATION'):
        base = base.filter(entry_type=entry_type)
    if ward_id:
        base = base.filter(ward_id=ward_id)
    if kit_id:
        base = base.filter(kiems_kit_id=kit_id)
    if date_from:
        try:
            base = base.filter(entry_date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            date_from = None
    if date_to:
        try:
            base = base.filter(entry_date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            date_to = None
    if search:
        base = base.filter(venue__icontains=search)

    # Count missing venues BEFORE the venue_status filter narrows the set further
    missing_count = base.filter(venue='').count()

    entries = base
    if venue_status == 'set':
        entries = entries.exclude(venue='')
    elif venue_status == 'missing':
        entries = entries.filter(venue='')

    entries = entries.order_by('-entry_date', 'ward__name')
    total_count = entries.count()

    # Duplicate venue/date detection across the WHOLE filtered set, not just the current page
    pairs = list(entries.exclude(venue='').values('id', 'venue', 'entry_date'))
    key_counts = Counter((p['venue'].strip().lower(), p['entry_date']) for p in pairs)
    conflict_ids = {
        p['id'] for p in pairs
        if key_counts[(p['venue'].strip().lower(), p['entry_date'])] > 1
    }

    paginator = Paginator(entries, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'wards': Ward.objects.all().order_by('name'),
        'kits': KIEMSKit.objects.all().order_by('kit_name'),
        'ward_selected': int(ward_id) if ward_id else None,
        'kit_selected': int(kit_id) if kit_id else None,
        'date_from': date_from,
        'date_to': date_to,
        'entry_type': entry_type,
        'venue_status': venue_status,
        'search': search,
        'total_count': total_count,
        'missing_count': missing_count,
        'conflict_ids': conflict_ids,
    }
    return render(request, 'superadmin/venue_management.html', context)


# ============================================================
# SHARED FILTER HELPER (ward / kit / date range)
# ============================================================
def _filtered_entries_qs(request):
    ward_id = request.GET.get('ward')
    kit_id = request.GET.get('kit')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    entry_type = request.GET.get('entry_type', '')   # '' = all types, matches venue_management now
    search = request.GET.get('search', '')

    entries = DailyKIEMSEntry.objects.select_related('kiems_kit', 'ward', 'vra', 'phase')

    if entry_type in ('VENUE', 'REGISTRATION'):
        entries = entries.filter(entry_type=entry_type)

    ward_obj = None
    kit_obj = None

    if ward_id:
        entries = entries.filter(ward_id=ward_id)
        ward_obj = Ward.objects.filter(id=ward_id).first()

    if kit_id:
        entries = entries.filter(kiems_kit_id=kit_id)
        kit_obj = KIEMSKit.objects.filter(id=kit_id).first()

    if date_from:
        try:
            entries = entries.filter(entry_date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            date_from = None

    if date_to:
        try:
            entries = entries.filter(entry_date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            date_to = None

    if search:
        entries = entries.filter(venue__icontains=search)

    entries = entries.order_by('entry_date', 'ward__name')

    return entries, {
        'ward_obj': ward_obj,
        'kit_obj': kit_obj,
        'date_from': date_from,
        'date_to': date_to,
        'entry_type': entry_type,
    }


# ============================================================
# DELETE / CLEAR VENUE
# ============================================================

@login_required
@user_passes_test(is_superadmin)
@require_POST
def delete_venue(request):
    """
    Clears the venue field on a single DailyKIEMSEntry. Does NOT delete the
    entry itself - registration counts, VRA, kit, and ward links are kept,
    only the venue text is removed and the entry goes back to "Pending".
    """
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': 'Invalid JSON data'}, status=400)

    entry_id = data.get('entry_id')
    if not entry_id:
        return JsonResponse({'ok': False, 'error': 'entry_id is required'}, status=400)

    try:
        entry = DailyKIEMSEntry.objects.get(id=entry_id)
    except (DailyKIEMSEntry.DoesNotExist, ValueError):
        return JsonResponse({'ok': False, 'error': 'Entry not found'}, status=404)

    entry.venue = ''
    entry.save(update_fields=['venue', 'updated_at'])

    return JsonResponse({
        'ok': True,
        'entry_id': entry.id,
        'message': 'Venue cleared.',
    })


# ============================================================
# VENUE / MOVEMENT SCHEDULE REPORT
# ============================================================

def _resolve_scope_names(filters, entries):
    """
    Work out what to print in the Ward / Kit meta boxes:
    the explicit filter if one was applied, otherwise the single
    value shared by every row, otherwise a generic label.
    """
    ward_obj = filters['ward_obj']
    kit_obj = filters['kit_obj']

    if ward_obj:
        ward_name = ward_obj.name
    else:
        ward_names = {e.ward.name for e in entries if e.ward}
        ward_name = ward_names.pop() if len(ward_names) == 1 else 'All Wards'

    if kit_obj:
        kit_no = f"{kit_obj.kit_name} ({kit_obj.serial_no})" if kit_obj.serial_no else kit_obj.kit_name
    else:
        kit_names = {
            f"{e.kiems_kit.kit_name} ({e.kiems_kit.serial_no})" if e.kiems_kit.serial_no else e.kiems_kit.kit_name
            for e in entries if e.kiems_kit
        }
        kit_no = kit_names.pop() if len(kit_names) == 1 else 'Multiple Kits'

    return ward_name, kit_no


from collections import defaultdict, Counter


def generate_venue_report_html(request):
    entries, filters = _filtered_entries_qs(request)
    entries = list(entries.exclude(venue=''))

    grouped_data = defaultdict(list)
    for e in entries:
        ward_name = e.ward.name if e.ward else "All Wards"
        kit_no = f"{e.kiems_kit.kit_name} ({e.kiems_kit.serial_no})" if e.kiems_kit.serial_no else e.kiems_kit.kit_name
        grouped_data[(ward_name, kit_no)].append(e)

    report_groups = []
    for (ward_name, kit_no), kit_entries in grouped_data.items():
        schedule = [{
            'date': e.entry_date.strftime('%d %b %Y'),
            'day': e.entry_date.strftime('%A'),
            'venue': e.venue,
            'time': '8:00 AM – 5:00 PM',
        } for e in kit_entries]
        report_groups.append({'ward': ward_name, 'kit_no': kit_no, 'schedule': schedule})

    html_string = render_to_string('superadmin/kitMovement_report.html', {
        'constituency': 'TURBO',
        'report_groups': report_groups,
        'has_entries': bool(report_groups),   # <-- new
        'date_from': filters['date_from'],
        'date_to': filters['date_to'],
        'generated_at': timezone.localtime().strftime('%d %b %Y, %H:%M'),
    })
    return HttpResponse(html_string)


@login_required
@user_passes_test(is_superadmin)
def generate_venue_report_pdf(request, as_attachment=False):
    """Generate the venue/movement PDF using PDF.co (same as generate_report_pdf)."""
    from django.conf import settings
    import requests

    try:
        if not hasattr(settings, 'PDF_CO_API_KEY') or not settings.PDF_CO_API_KEY:
            messages.error(request,
                           'PDF.co API key is not configured. Please add PDF_CO_API_KEY to your environment variables.')
            return redirect('superadmin:venue_management')

        html_response = generate_venue_report_html(request)
        html_string = html_response.content.decode('utf-8')

        timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')

        api_url = f"{getattr(settings, 'PDF_CO_API_URL', 'https://api.pdf.co/v1')}/pdf/convert/from/html"

        payload = json.dumps({
            "name": f"KIEMS_Movement_Notice_{timestamp}.pdf",
            "html": html_string,
            "margin": "20px",
            "paperSize": "Letter",
            "orientation": "Portrait",
            "printBackground": "true",
            "header": "",
            "footer": "",
            "async": False
        })

        headers = {
            'x-api-key': settings.PDF_CO_API_KEY,
            'Content-Type': 'application/json'
        }

        response_api = requests.post(api_url, headers=headers, data=payload, timeout=60)

        if response_api.status_code == 200:
            result = response_api.json()

            if result.get('error'):
                messages.error(request, f'PDF generation error: {result["error"]}')
                return redirect('superadmin:venue_management')

            if result.get('url'):
                pdf_response = requests.get(result['url'], timeout=30)
                if pdf_response.status_code == 200:
                    response = HttpResponse(pdf_response.content, content_type='application/pdf')
                    filename = f"KIEMS_Movement_Notice_{timestamp}.pdf"
                    disposition = 'attachment' if as_attachment else 'inline'
                    response['Content-Disposition'] = f'{disposition}; filename="{filename}"'
                    return response
                else:
                    messages.error(request, 'Failed to download generated PDF')
                    return redirect('superadmin:venue_management')
            elif result.get('file'):
                response = HttpResponse(result['file'], content_type='application/pdf')
                filename = f"KIEMS_Movement_Notice_{timestamp}.pdf"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            else:
                messages.error(request, 'Unexpected response from PDF service')
                return redirect('superadmin:venue_management')
        else:
            error_msg = response_api.json().get('error', 'Unknown error')
            messages.error(request, f'PDF generation service error: {error_msg}')
            return redirect('superadmin:venue_management')

    except requests.exceptions.Timeout:
        messages.error(request, 'PDF generation timed out. Please try again.')
        return redirect('superadmin:venue_management')
    except requests.exceptions.RequestException as e:
        messages.error(request, f'Network error: {str(e)}')
        return redirect('superadmin:venue_management')
    except Exception as e:
        messages.error(request, f'PDF generation failed: {str(e)}')
        return redirect('superadmin:venue_management')


@login_required
@user_passes_test(is_superadmin)
def generate_venue_report_pdf_fallback(request, as_attachment=False):
    """Local ReportLab fallback if PDF.co fails - mirrors generate_report_pdf_fallback."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
    except ImportError:
        messages.error(request, 'ReportLab is not available. Please install reportlab.')
        return redirect('superadmin:venue_management')

    entries, filters = _filtered_entries_qs(request)
    entries = list(entries.exclude(venue=''))
    ward_name, kit_no = _resolve_scope_names(filters, entries)

    timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(content_type='application/pdf')
    filename = f"KIEMS_Movement_Notice_{timestamp}.pdf"
    disposition = 'attachment' if as_attachment else 'inline'
    response['Content-Disposition'] = f'{disposition}; filename="{filename}"'

    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40,
    )

    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle(
        'Title', parent=styles['Title'], fontSize=16,
        textColor=colors.HexColor('#1B5E20'), alignment=TA_CENTER, spaceAfter=4,
    )
    story.append(Paragraph("NOTICE OF KIEMS KIT MOVEMENT SCHEDULE", title_style))

    subtitle_style = ParagraphStyle(
        'Subtitle', parent=styles['Normal'], fontSize=9,
        textColor=colors.HexColor('#6b7280'), alignment=TA_CENTER, spaceAfter=14,
    )
    story.append(Paragraph(f"Voter Registration &mdash; Turbo Constituency &bull; "
                           f"Generated: {timezone.localtime().strftime('%d %b %Y, %H:%M')}", subtitle_style))

    # Meta table (Constituency / Ward / Kit No)
    meta_data = [['Constituency', 'Ward', 'Kit No.'], ['Turbo', ward_name, kit_no]]
    meta_table = Table(meta_data, colWidths=[150, 180, 180])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 11),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#F5F5F5')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 14))

    # Body copy
    body_style = ParagraphStyle(
        'Body', parent=styles['Normal'], fontSize=10, alignment=TA_JUSTIFY, spaceAfter=14,
    )
    story.append(Paragraph(
        "This is to inform the general public, especially potential eligible applicants for registration "
        f"as voters, that the KIEMS Kit will be availed for registration of voters within "
        f"<b>{ward_name}</b> County Assembly Ward as follows:",
        body_style
    ))

    # Schedule table
    table_data = [['Date', 'Day', 'Venue', 'Time']]
    for row in entries:
        table_data.append([
            row.entry_date.strftime('%d %b %Y'),
            row.entry_date.strftime('%A'),
            row.venue,
            '8:00 AM - 5:00 PM',
        ])
    if len(table_data) == 1:
        table_data.append(['No schedule entries available', '', '', ''])

    schedule_table = Table(table_data, colWidths=[90, 90, 220, 110])
    schedule_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
    ]))
    story.append(schedule_table)
    story.append(Spacer(1, 18))

    # Requirements strip
    req_style = ParagraphStyle(
        'Req', parent=styles['Normal'], fontSize=9.5, textColor=colors.HexColor('#1B5E20'),
        backColor=colors.HexColor('#E8F5E9'), borderPadding=10, spaceAfter=30,
    )
    story.append(Paragraph(
        "<b>Requirements:</b> You must physically present yourself in person for biometric capture and be in "
        "possession of your original National ID or valid Kenyan Passport in order to register as a voter.",
        req_style
    ))

    # Signature block
    sig_style = ParagraphStyle('Sig', parent=styles['Normal'], fontSize=9, spaceBefore=30)
    sig_data = [
        ['Signed: ' + '.' * 30, 'Date: ' + '.' * 20, 'Official Stamp: ' + '.' * 15],
    ]
    sig_table = Table(sig_data, colWidths=[220, 150, 180])
    sig_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
    ]))
    story.append(sig_table)
    story.append(
        Paragraph('(RO / ARO / VRA)', ParagraphStyle('Caption', fontSize=8, textColor=colors.HexColor('#6b7280'))))

    doc.build(story)
    return response


@login_required
@user_passes_test(is_superadmin)
def generate_venue_report(request):
    """Try PDF.co first, fall back to ReportLab - same pattern as generate_report()."""
    try:
        return generate_venue_report_pdf(request, as_attachment=True)
    except Exception as e:
        messages.warning(request, f'PDF.co service failed, using fallback: {str(e)}')
        return generate_venue_report_pdf_fallback(request, as_attachment=True)


@login_required
@user_passes_test(is_superadmin)
def generate_venue_report_preview(request):
    """HTML preview of the movement notice, same pattern as generate_report_preview()."""
    return generate_venue_report_html(request)

@login_required
@user_passes_test(is_superadmin)
@require_GET
def entry_detail_api(request, pk):
    """Read-only entry details for the preview modal"""
    entry = get_object_or_404(
        DailyKIEMSEntry.objects.select_related('kiems_kit', 'phase', 'ward', 'vra', 'clerk'),
        pk=pk
    )
    return JsonResponse({
        'id': entry.id,
        'entry_type': entry.get_entry_type_display(),
        'date': entry.entry_date.strftime('%d %b %Y'),
        'kit': entry.kiems_kit.kit_name,
        'serial_no': entry.kiems_kit.serial_no,
        'phase': entry.phase.name,
        'ward': entry.ward.name,
        'vra': entry.vra.name,
        'clerk': entry.clerk.name if entry.clerk else '—',
        'venue': entry.venue or '—',
        'male': entry.registered_male,
        'female': entry.registered_female,
        'total_registered': entry.total_registered,
        'transferred': entry.total_transferred,
        'deleted': entry.total_updated,
        'uploaded': entry.uploaded,
        'edit_count': entry.edit_count,
        'created_at': entry.created_at.strftime('%d %b %Y, %H:%M'),
        'updated_at': entry.updated_at.strftime('%d %b %Y, %H:%M'),
        'office_updated_by': entry.office_updated_by or None,
        'office_updated_at': entry.office_updated_at.strftime('%d %b %Y, %H:%M') if entry.office_updated_at else None,
    })

@login_required
@user_passes_test(is_superadmin)
@require_GET
def get_kit_details(request):
    """Return a kit's ward, for auto-filling the entry form"""
    kit_id = request.GET.get('kit_id')
    if not kit_id:
        return JsonResponse({'error': 'No kit specified'}, status=400)
    try:
        kit = KIEMSKit.objects.select_related('ward').get(id=kit_id)
        return JsonResponse({'ward_id': kit.ward_id, 'ward_name': kit.ward.name})
    except KIEMSKit.DoesNotExist:
        return JsonResponse({'error': 'Kit not found'}, status=404)